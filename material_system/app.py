from __future__ import annotations

import cgi
import hashlib
import json
import mimetypes
import os
import re
import shutil
import sqlite3
import sys
import traceback
import uuid
from copy import copy
from datetime import datetime
from difflib import SequenceMatcher
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, unquote, urlparse

import openpyxl
import pandas as pd


ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data"
UPLOADS = DATA / "uploads"
OUTPUTS = DATA / "outputs"
STATIC = ROOT / "static"
DB = DATA / "material_versions.sqlite3"
DATE_FMT = "%Y-%m-%d %H:%M:%S"
APP_VERSION = "v1.0.27"
APP_VERSION_NOTE = "保護定版總表避免誤刪與降版"

MASTER_SHEETS = {"原物料", "包材"}
DEFAULT_KEY_FIELDS = ["來源", "化學物質", "CAS主號", "等級"]
COMPARE_FIELDS = ["化學物質", "CAS", "等級"]
SUBSTANCE_KEY_FIELDS = ["來源", "化學物質", "CAS主號", "等級"]
HEADER_LABELS = [
    "供應商Supplier",
    "供應商",
    "Supplier",
    "品名Item No.",
    "品名",
    "Item No.",
    "調查日期 Date",
    "調查日期",
    "Date",
    "填表人Name",
    "填表人",
    "公司印",
    "審核人",
    "Approve",
]
def now() -> str:
    return datetime.now().strftime(DATE_FMT)


def next_master_version() -> str:
    date_part = datetime.now().strftime("%Y%m%d")
    index = 1
    while True:
        version = f"MASTER-{date_part}-{index:03d}"
        if not (OUTPUTS / f"物料總表_{version}.xlsx").exists():
            return version
        index += 1


def ensure_dirs() -> None:
    for path in [DATA, UPLOADS, OUTPUTS, STATIC]:
        path.mkdir(parents=True, exist_ok=True)


def connect() -> sqlite3.Connection:
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    ensure_dirs()
    with connect() as conn:
        conn.executescript(
            """
            create table if not exists excel_files (
                id text primary key,
                kind text not null,
                original_name text not null,
                stored_path text not null,
                sha256 text not null,
                uploaded_at text not null,
                metadata_json text not null default '{}'
            );
            create table if not exists excel_versions (
                id text primary key,
                version_no text not null,
                file_id text not null,
                previous_version_no text,
                created_at text not null,
                operator text not null default '管理者',
                note text
            );
            create table if not exists comparison_jobs (
                id text primary key,
                master_file_id text not null,
                survey_file_id text not null,
                master_sheet text not null,
                material_name text,
                supplier_code text,
                category text,
                subcategory text,
                status text not null,
                created_at text not null,
                summary_json text not null
            );
            create table if not exists comparison_results (
                id text primary key,
                job_id text not null,
                diff_type text not null,
                composite_key text not null,
                old_json text not null,
                new_json text not null,
                changes_json text not null,
                action text not null default 'pending',
                note text,
                confirmed_at text
            );
            create table if not exists audit_logs (
                id text primary key,
                action text not null,
                target_id text,
                message text not null,
                created_at text not null
            );
            create table if not exists composite_key_settings (
                id integer primary key check (id = 1),
                fields_json text not null
            );
            create table if not exists app_state (
                key text primary key,
                value text not null
            );
            """,
        )
        conn.execute(
            "insert or ignore into composite_key_settings(id, fields_json) values (1, ?)",
            (json.dumps(DEFAULT_KEY_FIELDS, ensure_ascii=False),),
        )


def get_state(conn: sqlite3.Connection, key: str) -> str:
    row = conn.execute("select value from app_state where key=?", (key,)).fetchone()
    return row["value"] if row else ""


def set_state(conn: sqlite3.Connection, key: str, value: str) -> None:
    if value:
        conn.execute("insert or replace into app_state(key, value) values (?, ?)", (key, value))
    else:
        conn.execute("delete from app_state where key=?", (key,))


def version_state(conn: sqlite3.Connection) -> dict:
    notices = []
    latest_id = get_state(conn, "latest_master_file_id")
    working_id = get_state(conn, "working_master_file_id")
    latest = conn.execute("select id, kind, original_name, uploaded_at from excel_files where id=?", (latest_id,)).fetchone() if latest_id else None
    working = conn.execute("select id, kind, original_name, uploaded_at from excel_files where id=?", (working_id,)).fetchone() if working_id else None
    if latest_id and not latest:
        set_state(conn, "latest_master_file_id", "")
        recovered = conn.execute(
            """
            select f.id, f.kind, f.original_name, f.uploaded_at
            from excel_versions v
            join excel_files f on f.id = v.file_id
            where instr(coalesce(v.note, ''), '已定為最新版本') > 0
               or instr(coalesce(v.note, ''), '目前最新總表') > 0
            order by v.created_at desc
            limit 1
            """
        ).fetchone()
        if recovered:
            set_state(conn, "latest_master_file_id", recovered["id"])
            latest = recovered
            notices.append({
                "type": "latest_master_recovered",
                "message": f"系統偵測到原本記錄的最新總表不存在，已依版本紀錄自動恢復為：{recovered['original_name']}。請確認是否符合預期。",
                "file_id": recovered["id"],
            })
        else:
            notices.append({
                "type": "latest_master_missing",
                "message": "系統偵測到原本記錄的最新總表不存在，且版本紀錄中找不到可自動恢復的總表。請重新指定最新總表。",
            })
    if working_id and not working:
        set_state(conn, "working_master_file_id", "")
        notices.append({
            "type": "working_master_missing",
            "message": "系統偵測到本輪暫存更動版不存在，已清除暫存狀態。",
        })
    return {
        "latest_master": dict(latest) if latest else None,
        "working_master": dict(working) if working else None,
        "notices": notices,
    }


def log(action: str, message: str, target_id: str | None = None) -> None:
    with connect() as conn:
        conn.execute(
            "insert into audit_logs values (?, ?, ?, ?, ?)",
            (str(uuid.uuid4()), action, target_id, message, now()),
        )


def file_hash(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def clean_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    text = str(value).replace("\u3000", " ").replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text).strip()
    return text


def normalize_key_value(field: str, value) -> str:
    text = clean_text(value)
    if field in {"CAS", "供應商代號", "物料編號"}:
        text = re.sub(r"\s+", "", text)
    return text.upper()


def loose_token(value) -> str:
    text = normalize_key_value("", value)
    return re.sub(r"[^0-9A-Z\u4e00-\u9fff]+", "", text)


def excel_col_name(col: int) -> str:
    name = ""
    while col:
        col, rem = divmod(col - 1, 26)
        name = chr(65 + rem) + name
    return name


def cas_primary(value) -> str:
    text = clean_text(value)
    match = re.search(r"\b\d{2,7}-\d{2}-\d\b", text)
    if match:
        return match.group(0)
    if text in {"-", "同上"}:
        return ""
    return text


def compare_score(needle: str, candidate: str, category: str, candidate_category: str) -> tuple[int, list[str]]:
    reasons = []
    needle_token = loose_token(needle)
    candidate_token = loose_token(candidate)
    if not needle_token or not candidate_token:
        return 0, reasons
    ratio = SequenceMatcher(None, needle_token, candidate_token).ratio()
    score = int(ratio * 70)
    if needle_token == candidate_token:
        score = 100
        reasons.append("去符號後相同")
    elif needle_token in candidate_token or candidate_token in needle_token:
        score = max(score, 86)
        reasons.append("品名包含")
    elif score >= 55:
        reasons.append("品名相似")
    if category and normalize_key_value("化學物質類別", category) == normalize_key_value("化學物質類別", candidate_category):
        score += 8
        reasons.append("同類別")
    return min(score, 100), reasons


def normalize_compare_value(value) -> str:
    text = clean_text(value)
    lowered = text.lower()
    if lowered in {"", "nan", "none"}:
        return ""
    if lowered in {"no contain", "nd", "無no", "無"}:
        return "No contain"
    return text


def has_lookup_content(value: str) -> bool:
    normalized = normalize_compare_value(value)
    return bool(normalized) and normalized.lower() != "no contain"


def make_key(record: dict, fields: list[str]) -> str:
    return " | ".join(normalize_key_value(field, record.get(field)) for field in fields)


def enrich_substance_key(record: dict) -> dict:
    enriched = dict(record)
    enriched["CAS主號"] = cas_primary(record.get("CAS"))
    return enriched


def workbook_info(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        sheets = []
        for ws in wb.worksheets:
            preview = []
            max_row = ws.max_row or 0
            max_column = ws.max_column or 0
            for row in ws.iter_rows(min_row=1, max_row=min(max_row, 20), values_only=True):
                preview.append([clean_text(v) for v in row[: min(max_column, 12)]])
            sheets.append(
                {
                    "name": ws.title,
                    "rows": max_row,
                    "columns": max_column,
                    "hidden": ws.sheet_state != "visible",
                    "preview": preview,
                }
            )
        return {"sheets": sheets}
    finally:
        wb.close()


def resolve_sheet_name(wb, requested: str) -> str:
    requested_clean = clean_text(requested)
    if requested_clean in wb.sheetnames:
        return requested_clean
    requested_token = loose_token(requested_clean)
    for name in wb.sheetnames:
        if loose_token(name) == requested_token:
            return name
    for name in wb.sheetnames:
        if requested_token and (requested_token in loose_token(name) or loose_token(name) in requested_token):
            return name
    if any(loose_token(name) == loose_token("原物料") for name in wb.sheetnames):
        return next(name for name in wb.sheetnames if loose_token(name) == loose_token("原物料"))
    return wb.sheetnames[0]


def looks_like_supplier_code(value: str) -> bool:
    value = clean_text(value)
    if not value:
        return False
    has_cjk = any("\u4e00" <= char <= "\u9fff" for char in value)
    has_digit = any(char.isdigit() for char in value)
    return has_digit and not has_cjk


def package_layout(ws) -> dict:
    material_row = 5
    for row in range(3, min(ws.max_row, 8) + 1):
        left_header = f"{clean_text(ws.cell(row, 2).value)} {clean_text(ws.cell(row, 3).value)} {clean_text(ws.cell(row, 4).value)}".lower()
        if ("chemical" in left_header or "化學物質" in left_header) and ("level" in left_header or "等級" in left_header):
            material_row = row
            break
    supplier_code_row = None
    supplier_row = None
    for row in range(1, material_row):
        label = clean_text(ws.cell(row, 4).value)
        if "供應商代號" in label:
            supplier_code_row = row
        elif "供應商" in label:
            supplier_row = row
    if supplier_row is None and material_row > 1:
        supplier_row = material_row - 1
    return {
        "material_start": 5,
        "material_row": material_row,
        "data_start": material_row + 1,
        "supplier_code_row": supplier_code_row,
        "supplier_row": supplier_row,
    }


def package_supplier_code(ws, col: int, layout: dict | None = None) -> str:
    layout = layout or package_layout(ws)
    row = layout.get("supplier_code_row")
    return clean_text(ws.cell(row, col).value) if row else ""


def package_supplier_name(ws, col: int, layout: dict | None = None) -> str:
    layout = layout or package_layout(ws)
    row = layout.get("supplier_row")
    return clean_text(ws.cell(row, col).value) if row else ""


def package_supplier_matches(ws, col: int, supplier_identifier: str, layout: dict | None = None) -> bool:
    supplier_identifier = clean_text(supplier_identifier)
    if not supplier_identifier:
        return True
    layout = layout or package_layout(ws)
    wanted_code = normalize_key_value("供應商代號", supplier_identifier)
    wanted_name = normalize_key_value("供應商", supplier_identifier)
    candidates = [
        normalize_key_value("供應商代號", package_supplier_code(ws, col, layout)),
        normalize_key_value("供應商", package_supplier_name(ws, col, layout)),
    ]
    return wanted_code in candidates or wanted_name in candidates


def parse_master(path: Path, sheet_name: str | None = None) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    records: list[dict] = []
    target_sheets = [sheet_name] if sheet_name else [s for s in wb.sheetnames if s in MASTER_SHEETS]
    for name in target_sheets:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        if name == "原物料":
            material_start, data_start = 6, 9
            for col in range(material_start, ws.max_column + 1):
                material = clean_text(ws.cell(5, col).value)
                if not material:
                    continue
                meta = {
                    "來源": name,
                    "物料名稱": material,
                    "化學物質類別": clean_text(ws.cell(6, col).value),
                    "細項類別": "",
                    "供應商代號": clean_text(ws.cell(7, col).value),
                    "供應商": clean_text(ws.cell(8, col).value),
                    "excel_row": None,
                    "excel_col": col,
                }
                for row in range(data_start, ws.max_row + 1):
                    substance = clean_text(ws.cell(row, 3).value)
                    if not substance:
                        continue
                    records.append(
                        {
                            **meta,
                            "excel_row": row,
                            "項次": clean_text(ws.cell(row, 2).value),
                            "化學物質": substance,
                            "CAS": clean_text(ws.cell(row, 4).value),
                            "允許濃度": clean_text(ws.cell(row, 4).value),
                            "等級": clean_text(ws.cell(row, 5).value),
                            "含有值": normalize_compare_value(ws.cell(row, col).value),
                        }
                    )
        elif name == "包材":
            layout = package_layout(ws)
            material_start, data_start = layout["material_start"], layout["data_start"]
            for col in range(material_start, ws.max_column + 1):
                material = clean_text(ws.cell(layout["material_row"], col).value)
                if not material:
                    continue
                meta = {
                    "來源": name,
                    "物料名稱": material,
                    "化學物質類別": "包材",
                    "細項類別": "",
                    "供應商代號": package_supplier_code(ws, col, layout),
                    "供應商": package_supplier_name(ws, col, layout),
                    "excel_row": None,
                    "excel_col": col,
                }
                for row in range(data_start, ws.max_row + 1):
                    substance = clean_text(ws.cell(row, 2).value)
                    if not substance:
                        continue
                    records.append(
                        {
                            **meta,
                            "excel_row": row,
                            "項次": clean_text(ws.cell(row, 1).value),
                            "化學物質": substance,
                            "CAS": clean_text(ws.cell(row, 3).value),
                            "允許濃度": clean_text(ws.cell(row, 3).value),
                            "等級": clean_text(ws.cell(row, 4).value),
                            "含有值": normalize_compare_value(ws.cell(row, col).value),
                        }
                    )
    wb.close()
    return records


def parse_master_material(path: Path, sheet_name: str, material_name: str, supplier_code: str = "") -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    target_col = None
    if sheet_name == "原物料":
        for col in range(6, ws.max_column + 1):
            same_material = normalize_key_value("物料名稱", ws.cell(5, col).value) == normalize_key_value("物料名稱", material_name)
            same_supplier = not supplier_code or normalize_key_value("供應商代號", ws.cell(7, col).value) == normalize_key_value("供應商代號", supplier_code)
            if same_material and same_supplier:
                target_col = col
                break
        if target_col is None:
            return []
        meta = {
            "來源": sheet_name,
            "物料名稱": clean_text(ws.cell(5, target_col).value),
            "化學物質類別": clean_text(ws.cell(6, target_col).value),
            "細項類別": "",
            "供應商代號": clean_text(ws.cell(7, target_col).value),
            "供應商": clean_text(ws.cell(8, target_col).value),
            "excel_col": target_col,
        }
        return [
            {
                **meta,
                "excel_row": row,
                "項次": clean_text(ws.cell(row, 2).value),
                "化學物質": clean_text(ws.cell(row, 3).value),
                "CAS": clean_text(ws.cell(row, 4).value),
                "允許濃度": clean_text(ws.cell(row, 4).value),
                "等級": clean_text(ws.cell(row, 5).value),
                "含有值": normalize_compare_value(ws.cell(row, target_col).value),
            }
            for row in range(9, ws.max_row + 1)
            if clean_text(ws.cell(row, 3).value)
        ]
    if sheet_name == "包材":
        layout = package_layout(ws)
        for col in range(layout["material_start"], ws.max_column + 1):
            same_material = normalize_key_value("物料名稱", ws.cell(layout["material_row"], col).value) == normalize_key_value("物料名稱", material_name)
            same_supplier = package_supplier_matches(ws, col, supplier_code, layout)
            if same_material and same_supplier:
                target_col = col
                break
        if target_col is None:
            return []
        meta = {
            "來源": sheet_name,
            "物料名稱": clean_text(ws.cell(layout["material_row"], target_col).value),
            "化學物質類別": "包材",
            "細項類別": "",
            "供應商代號": package_supplier_code(ws, target_col, layout),
            "供應商": package_supplier_name(ws, target_col, layout),
            "excel_col": target_col,
        }
        return [
            {
                **meta,
                "excel_row": row,
                "項次": clean_text(ws.cell(row, 1).value),
                "化學物質": clean_text(ws.cell(row, 2).value),
                "CAS": clean_text(ws.cell(row, 3).value),
                "允許濃度": clean_text(ws.cell(row, 3).value),
                "等級": clean_text(ws.cell(row, 4).value),
                "含有值": normalize_compare_value(ws.cell(row, target_col).value),
            }
            for row in range(layout["data_start"], ws.max_row + 1)
            if clean_text(ws.cell(row, 2).value)
        ]
    return []


def parse_master_column(path: Path, sheet_name: str, target_col: int) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    sheet_name = resolve_sheet_name(wb, sheet_name)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    if target_col < 1 or target_col > ws.max_column:
        return []
    if sheet_name == "原物料":
        meta = {
            "來源": sheet_name,
            "物料名稱": clean_text(ws.cell(5, target_col).value),
            "化學物質類別": clean_text(ws.cell(6, target_col).value),
            "細項類別": "",
            "供應商代號": clean_text(ws.cell(7, target_col).value),
            "供應商": clean_text(ws.cell(8, target_col).value),
            "excel_col": target_col,
        }
        records = []
        for row_index, row in enumerate(ws.iter_rows(min_row=9, max_row=ws.max_row, min_col=1, max_col=target_col), start=9):
            substance = clean_text(row[2].value)
            if not substance:
                continue
            target_value = row[target_col - 1].value
            records.append(
                enrich_substance_key(
                    {
                        **meta,
                        "excel_row": row_index,
                        "項次": clean_text(row[1].value),
                        "化學物質": substance,
                        "CAS": clean_text(row[3].value),
                        "允許濃度": clean_text(row[3].value),
                        "等級": clean_text(row[4].value),
                        "含有值": normalize_compare_value(target_value),
                        "公式儲存格": isinstance(target_value, str) and target_value.startswith("="),
                    }
                )
            )
        return records
    if sheet_name == "包材":
        layout = package_layout(ws)
        meta = {
            "來源": sheet_name,
            "物料名稱": clean_text(ws.cell(layout["material_row"], target_col).value),
            "化學物質類別": "包材",
            "細項類別": "",
            "供應商代號": package_supplier_code(ws, target_col, layout),
            "供應商": package_supplier_name(ws, target_col, layout),
            "excel_col": target_col,
        }
        records = []
        for row_index, row in enumerate(ws.iter_rows(min_row=layout["data_start"], max_row=ws.max_row, min_col=1, max_col=target_col), start=layout["data_start"]):
            substance = clean_text(row[1].value)
            if not substance:
                continue
            target_value = row[target_col - 1].value
            records.append(
                enrich_substance_key(
                    {
                        **meta,
                        "excel_row": row_index,
                        "項次": clean_text(row[0].value),
                        "化學物質": substance,
                        "CAS": clean_text(row[2].value),
                        "允許濃度": clean_text(row[2].value),
                        "等級": clean_text(row[3].value),
                        "含有值": normalize_compare_value(target_value),
                        "公式儲存格": isinstance(target_value, str) and target_value.startswith("="),
                    }
                )
            )
        return records
    return []


def parse_master_substances(path: Path, sheet_name: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    sheet_name = resolve_sheet_name(wb, sheet_name)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    records = []
    if sheet_name == "原物料":
        for row_index, row in enumerate(ws.iter_rows(min_row=9, max_row=ws.max_row, min_col=1, max_col=5), start=9):
            substance = clean_text(row[2].value)
            if not substance:
                continue
            records.append(
                enrich_substance_key(
                    {
                        "來源": sheet_name,
                        "物料名稱": "",
                        "物料編號": "",
                        "化學物質類別": "",
                        "細項類別": "",
                        "供應商代號": "",
                        "供應商": "",
                        "excel_row": row_index,
                        "excel_col": None,
                        "項次": clean_text(row[1].value),
                        "化學物質": substance,
                        "CAS": clean_text(row[3].value),
                        "允許濃度": "",
                        "等級": clean_text(row[4].value),
                        "含有值": "",
                    }
                )
            )
    elif sheet_name == "包材":
        layout = package_layout(ws)
        for row_index, row in enumerate(ws.iter_rows(min_row=layout["data_start"], max_row=ws.max_row, min_col=1, max_col=4), start=layout["data_start"]):
            substance = clean_text(row[1].value)
            if not substance:
                continue
            records.append(
                enrich_substance_key(
                    {
                        "來源": sheet_name,
                        "物料名稱": "",
                        "物料編號": "",
                        "化學物質類別": "包材",
                        "細項類別": "",
                        "供應商代號": "",
                        "供應商": "",
                        "excel_row": row_index,
                        "excel_col": None,
                        "項次": clean_text(row[0].value),
                        "化學物質": substance,
                        "CAS": clean_text(row[2].value),
                        "允許濃度": "",
                        "等級": clean_text(row[3].value),
                        "含有值": "",
                    }
                )
            )
    wb.close()
    return records


def material_candidates(path: Path, sheet_name: str, material_name: str, supplier_code: str, category: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    sheet_name = resolve_sheet_name(wb, sheet_name)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    candidates = []
    if sheet_name == "原物料":
        for col in range(6, ws.max_column + 1):
            candidate_supplier_code = clean_text(ws.cell(7, col).value)
            if supplier_code and normalize_key_value("供應商代號", candidate_supplier_code) != normalize_key_value("供應商代號", supplier_code):
                continue
            candidate_material = clean_text(ws.cell(5, col).value)
            candidate_category = clean_text(ws.cell(6, col).value)
            score, reasons = compare_score(material_name, candidate_material, category, candidate_category)
            if supplier_code:
                score += 10
                reasons.append("同供應商代號")
            if score >= 50:
                candidates.append(
                    {
                        "sheet": sheet_name,
                        "column": col,
                        "column_name": excel_col_name(col),
                        "material_name": candidate_material,
                        "supplier_code": candidate_supplier_code,
                        "supplier": clean_text(ws.cell(8, col).value),
                        "category": candidate_category,
                        "subcategory": "",
                        "score": min(score, 100),
                        "reasons": reasons,
                    }
                )
    elif sheet_name == "包材":
        layout = package_layout(ws)
        for col in range(layout["material_start"], ws.max_column + 1):
            candidate_supplier_code = package_supplier_code(ws, col, layout)
            candidate_supplier = package_supplier_name(ws, col, layout)
            if supplier_code and not package_supplier_matches(ws, col, supplier_code, layout):
                continue
            candidate_material = clean_text(ws.cell(layout["material_row"], col).value)
            score, reasons = compare_score(material_name, candidate_material, category, "包材")
            if supplier_code:
                score += 10
                reasons.append("同供應商")
            if score >= 50:
                candidates.append(
                    {
                        "sheet": sheet_name,
                        "column": col,
                        "column_name": excel_col_name(col),
                        "material_name": candidate_material,
                        "supplier_code": candidate_supplier_code,
                        "supplier": candidate_supplier,
                        "category": "包材",
                        "subcategory": "",
                        "score": min(score, 100),
                        "reasons": reasons,
                    }
                )
    wb.close()
    return sorted(candidates, key=lambda x: (-x["score"], x["column"]))


def material_columns(path: Path, sheet_name: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    sheet_name = resolve_sheet_name(wb, sheet_name)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    columns = []
    if sheet_name == "原物料":
        for col in range(6, ws.max_column + 1):
            material = clean_text(ws.cell(5, col).value)
            if not material:
                continue
            supplier_code = clean_text(ws.cell(7, col).value)
            supplier = clean_text(ws.cell(8, col).value)
            category = clean_text(ws.cell(6, col).value)
            columns.append({
                "sheet": sheet_name,
                "column": col,
                "column_name": excel_col_name(col),
                "material_name": material,
                "supplier_code": supplier_code,
                "supplier": supplier,
                "category": category,
                "subcategory": "",
                "label": f"{excel_col_name(col)}欄｜{material}｜{supplier_code}｜{supplier}｜{category}",
            })
    elif sheet_name == "包材":
        layout = package_layout(ws)
        for col in range(layout["material_start"], ws.max_column + 1):
            material = clean_text(ws.cell(layout["material_row"], col).value)
            if not material:
                continue
            supplier_code = package_supplier_code(ws, col, layout)
            supplier = package_supplier_name(ws, col, layout)
            columns.append({
                "sheet": sheet_name,
                "column": col,
                "column_name": excel_col_name(col),
                "material_name": material,
                "supplier_code": supplier_code,
                "supplier": supplier,
                "category": "包材",
                "subcategory": "",
                "label": f"{excel_col_name(col)}欄｜{material}｜{supplier_code or supplier}｜包材",
            })
    wb.close()
    return columns


def query_master_details(path: Path, sheet_name: str, filters: dict) -> dict:
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    sheet_name = resolve_sheet_name(wb, sheet_name)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {"sheet": sheet_name, "materials": [], "rows": []}
    ws = wb[sheet_name]
    supplier_code_q = loose_token(filters.get("supplier_code"))
    supplier_q = loose_token(filters.get("supplier"))
    material_q = loose_token(filters.get("material_name"))
    category_q = loose_token(filters.get("category"))
    substance_q = loose_token(filters.get("substance"))
    include_empty = bool(filters.get("include_empty"))
    page = max(1, int(filters.get("page") or 1))
    page_size = min(500, max(50, int(filters.get("page_size") or 200)))

    materials = []
    rows = []
    if sheet_name == "原物料":
        material_start, data_start = 6, 9
        indexes = {"item": 2, "substance": 3, "cas": 4, "level": 5}
        for col in range(material_start, ws.max_column + 1):
            material = clean_text(ws.cell(5, col).value)
            if not material:
                continue
            category = clean_text(ws.cell(6, col).value)
            supplier_code = clean_text(ws.cell(7, col).value)
            supplier = clean_text(ws.cell(8, col).value)
            if supplier_code_q and supplier_code_q not in loose_token(supplier_code):
                continue
            if supplier_q and supplier_q not in loose_token(supplier):
                continue
            if material_q and material_q not in loose_token(material):
                continue
            if category_q and category_q not in loose_token(category):
                continue
            materials.append(
                {
                    "sheet": sheet_name,
                    "column": col,
                    "column_name": excel_col_name(col),
                    "material_name": material,
                    "category": category,
                    "subcategory": "",
                    "supplier_code": supplier_code,
                    "supplier": supplier,
                }
            )
            for row in range(data_start, ws.max_row + 1):
                substance = clean_text(ws.cell(row, indexes["substance"]).value)
                if not substance:
                    continue
                if substance_q and substance_q not in loose_token(substance):
                    continue
                content = normalize_compare_value(ws.cell(row, col).value)
                if not include_empty and not has_lookup_content(content):
                    continue
                rows.append(
                    {
                        "sheet": sheet_name,
                        "column_name": excel_col_name(col),
                        "material_name": material,
                        "category": category,
                        "subcategory": "",
                        "supplier_code": supplier_code,
                        "supplier": supplier,
                        "item": clean_text(ws.cell(row, indexes["item"]).value),
                        "substance": substance,
                        "cas": clean_text(ws.cell(row, indexes["cas"]).value),
                        "cas_primary": cas_primary(ws.cell(row, indexes["cas"]).value),
                        "level": clean_text(ws.cell(row, indexes["level"]).value),
                        "content": content,
                        "excel_row": row,
                        "excel_col": col,
                    }
                )
    elif sheet_name == "包材":
        layout = package_layout(ws)
        material_start, data_start = layout["material_start"], layout["data_start"]
        indexes = {"item": 1, "substance": 2, "cas": 3, "level": 4}
        for col in range(material_start, ws.max_column + 1):
            material = clean_text(ws.cell(layout["material_row"], col).value)
            if not material:
                continue
            category = "包材"
            supplier_code = package_supplier_code(ws, col, layout)
            supplier = package_supplier_name(ws, col, layout)
            if supplier_code_q and supplier_code_q not in loose_token(supplier_code):
                continue
            if supplier_q and supplier_q not in loose_token(supplier):
                continue
            if material_q and material_q not in loose_token(material):
                continue
            if category_q and category_q not in loose_token(category):
                continue
            materials.append(
                {
                    "sheet": sheet_name,
                    "column": col,
                    "column_name": excel_col_name(col),
                    "material_name": material,
                    "category": category,
                    "subcategory": "",
                    "supplier_code": supplier_code,
                    "supplier": supplier,
                }
            )
            for row in range(data_start, ws.max_row + 1):
                substance = clean_text(ws.cell(row, indexes["substance"]).value)
                if not substance:
                    continue
                if substance_q and substance_q not in loose_token(substance):
                    continue
                content = normalize_compare_value(ws.cell(row, col).value)
                if not include_empty and not has_lookup_content(content):
                    continue
                rows.append(
                    {
                        "sheet": sheet_name,
                        "column_name": excel_col_name(col),
                        "material_name": material,
                        "category": category,
                        "subcategory": "",
                        "supplier_code": supplier_code,
                        "supplier": supplier,
                        "item": clean_text(ws.cell(row, indexes["item"]).value),
                        "substance": substance,
                        "cas": clean_text(ws.cell(row, indexes["cas"]).value),
                        "cas_primary": cas_primary(ws.cell(row, indexes["cas"]).value),
                        "level": clean_text(ws.cell(row, indexes["level"]).value),
                        "content": content,
                        "excel_row": row,
                        "excel_col": col,
                    }
                )
    total_rows = len(rows)
    start = (page - 1) * page_size
    wb.close()
    return {
        "sheet": sheet_name,
        "materials": materials,
        "rows": rows[start:start + page_size],
        "total_rows": total_rows,
        "page": page,
        "page_size": page_size,
    }


def master_filter_options(path: Path, sheet_name: str) -> dict:
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    sheet_name = resolve_sheet_name(wb, sheet_name)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {"sheet": sheet_name, "supplier_codes": [], "suppliers": [], "materials": [], "categories": [], "substances": []}
    ws = wb[sheet_name]
    supplier_codes: set[str] = set()
    suppliers: set[str] = set()
    materials: set[str] = set()
    categories: set[str] = set()
    substances: set[str] = set()
    if sheet_name == "原物料":
        for col in range(6, ws.max_column + 1):
            material = clean_text(ws.cell(5, col).value)
            if not material:
                continue
            materials.add(material)
            category = clean_text(ws.cell(6, col).value)
            supplier_code = clean_text(ws.cell(7, col).value)
            supplier = clean_text(ws.cell(8, col).value)
            if category:
                categories.add(category)
            if supplier_code:
                supplier_codes.add(supplier_code)
            if supplier:
                suppliers.add(supplier)
        for row in range(9, ws.max_row + 1):
            substance = clean_text(ws.cell(row, 3).value)
            if substance:
                substances.add(substance)
    elif sheet_name == "包材":
        categories.add("包材")
        layout = package_layout(ws)
        for col in range(layout["material_start"], ws.max_column + 1):
            material = clean_text(ws.cell(layout["material_row"], col).value)
            if not material:
                continue
            materials.add(material)
            supplier_code = package_supplier_code(ws, col, layout)
            supplier = package_supplier_name(ws, col, layout)
            if supplier_code:
                supplier_codes.add(supplier_code)
            if supplier:
                suppliers.add(supplier)
        for row in range(layout["data_start"], ws.max_row + 1):
            substance = clean_text(ws.cell(row, 2).value)
            if substance:
                substances.add(substance)
    wb.close()
    return {
        "sheet": sheet_name,
        "supplier_codes": sorted(supplier_codes, key=loose_token),
        "suppliers": sorted(suppliers, key=loose_token),
        "materials": sorted(materials, key=loose_token),
        "categories": sorted(categories, key=loose_token),
        "substances": sorted(substances, key=loose_token),
    }


def parse_meta_value(text: str, label: str) -> str:
    match = re.search(label + r"\s*[:：]\s*([^\n﹙]+)", text)
    return clean_text(match.group(1)) if match else ""


def parse_labeled_value(ws, labels: list[str], max_row: int = 6, max_col: int = 10) -> str:
    label_tokens = [loose_token(label) for label in labels]
    all_label_tokens = [loose_token(label) for label in HEADER_LABELS]
    for row in range(1, min(ws.max_row, max_row) + 1):
        for col in range(1, min(ws.max_column, max_col) + 1):
            value = clean_text(ws.cell(row, col).value)
            if not value:
                continue
            value_token = loose_token(value)
            if not any(token and token in value_token for token in label_tokens):
                continue
            for label in labels:
                parsed = parse_meta_value(value, label)
                if parsed:
                    return parsed
            parts = re.split(r"[:：]", value, maxsplit=1)
            if len(parts) == 2:
                tail = clean_text(re.sub(r"﹙.*", "", parts[1]))
                if tail:
                    return tail
            for next_col in range(col + 1, min(ws.max_column, max_col) + 1):
                neighbor = clean_text(ws.cell(row, next_col).value)
                neighbor_token = loose_token(neighbor)
                if neighbor and not any(token and token in neighbor_token for token in all_label_tokens):
                    return neighbor
    return ""


def infer_material_from_filename(path: Path) -> str:
    stem = path.stem
    stem = re.sub(r"^[0-9a-fA-F-]{32,36}_", "", stem)
    text = re.sub(r"WI-QA-\d+-\d+[-_ ]*", "", stem, flags=re.I)
    text = re.sub(r"有害物質調查表", "", text)
    text = re.sub(r"Ver\s*\d+", "", text, flags=re.I)
    text = re.sub(r"\(\d+\)", "", text)
    text = re.sub(r"\b\d{4,8}\b", "", text)
    text = clean_text(text)
    return text if len(loose_token(text)) >= 3 else ""


def parse_survey_header(path: Path) -> tuple[object, object, dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    sheet = "有害物質調查表" if "有害物質調查表" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]
    supplier = parse_labeled_value(ws, ["供應商Supplier", "供應商", "Supplier"])
    material = parse_labeled_value(ws, ["品名Item No.", "品名", "Item No."])
    material_source = "調查表表頭"
    if not material:
        material = infer_material_from_filename(path)
        material_source = "檔名推測" if material else "未解析"
    survey_date = parse_labeled_value(ws, ["調查日期 Date", "調查日期", "Date"])
    meta = {
        "sheet": sheet,
        "supplier": supplier,
        "material_name": material,
        "material_source": material_source,
        "survey_date": survey_date,
    }
    return wb, ws, meta


def parse_survey(path: Path, supplier_code: str, category: str, subcategory: str, version: str) -> tuple[list[dict], dict]:
    _wb, ws, meta = parse_survey_header(path)
    supplier = meta["supplier"]
    material = meta["material_name"]
    records = []
    warnings = []
    for row_index, row in enumerate(ws.iter_rows(min_row=9, max_row=ws.max_row, min_col=1, max_col=8), start=9):
        substance = clean_text(row[1].value)
        if not substance or substance.lower().startswith("level"):
            continue
        yes_no = clean_text(row[5].value)
        content = clean_text(row[7].value)
        issue = ""
        if "有" in yes_no and not content:
            warnings.append(f"第 {row_index} 列標示有含有，但含量空白；此版先列為備註，不阻擋細項比對")
        value = content if content else ("No contain" if "無" in yes_no else "")
        records.append(
            enrich_substance_key(
                {
                    "來源": "原物料" if category != "包材" else "包材",
                    "物料名稱": material,
                    "物料編號": material,
                    "化學物質類別": category,
                    "細項類別": subcategory,
                    "供應商代號": supplier_code,
                    "供應商": supplier,
                    "調查表版本": version,
                    "項次": clean_text(row[0].value),
                    "化學物質": substance,
                    "CAS": clean_text(row[2].value),
                    "允許濃度": clean_text(row[4].value),
                    "等級": clean_text(row[3].value),
                    "含有值": normalize_compare_value(value),
                    "資料問題": issue,
                    "excel_row": row_index,
                    "excel_col": None,
                }
            )
        )
    meta.update({"warnings": warnings, "rows": len(records)})
    _wb.close()
    return records, meta


def survey_update_value(record: dict, sheet_name: str) -> str:
    if sheet_name == "包材":
        return normalize_compare_value(record.get("允許濃度"))
    return normalize_compare_value(record.get("含有值"))


def effective_survey_update_value(record: dict, sheet_name: str) -> str:
    value = survey_update_value(record, sheet_name)
    return value or "No contain"


def values_differ(field: str, old_value, new_value) -> bool:
    return normalize_compare_value(old_value) != normalize_compare_value(new_value)


def row_index_by_substance(records: list[dict]) -> dict[str, dict]:
    indexed: dict[str, dict] = {}
    for record in records:
        key = make_key(record, SUBSTANCE_KEY_FIELDS)
        if key not in indexed:
            indexed[key] = record
    return indexed


def compare_batch_records(master_path: Path, sheet_name: str, new_records: list[dict], old_records: list[dict] | None = None) -> tuple[list[dict], dict]:
    wb = openpyxl.load_workbook(master_path, read_only=False, data_only=False)
    sheet_name = resolve_sheet_name(wb, sheet_name)
    ws = wb[sheet_name]
    if old_records is None:
        old_records = parse_master_substances(master_path, sheet_name)
    old_by_key = row_index_by_substance(old_records)
    results = []
    summary = {"added": 0, "removed": 0, "modified": 0, "same": 0, "conflict": 0}
    seen: dict[str, dict] = {}
    material_col_cache: dict[tuple[str, str], int | None] = {}
    for new in new_records:
        row_key = make_key(new, SUBSTANCE_KEY_FIELDS)
        target_key = make_key(new, [*SUBSTANCE_KEY_FIELDS, "供應商代號", "物料名稱"])
        new_value = effective_survey_update_value(new, sheet_name)
        changes = []
        if target_key in seen and effective_survey_update_value(seen[target_key], sheet_name) != new_value:
            old = seen[target_key]
            changes.append({"field": "同批重複資料", "old": effective_survey_update_value(old, sheet_name), "new": new_value})
            diff_type = "modified"
        else:
            seen[target_key] = new
            old = dict(old_by_key.get(row_key, {}))
            if old:
                material = new.get("物料名稱", "")
                supplier_code = new.get("供應商代號", "")
                col_key = (normalize_key_value("物料名稱", material), normalize_key_value("供應商代號", supplier_code))
                if col_key not in material_col_cache:
                    material_col_cache[col_key] = find_material_column(ws, sheet_name, material, supplier_code)
                col = material_col_cache[col_key]
                old_value = normalize_compare_value(ws.cell(old["excel_row"], col).value) if col else "No contain"
                old["含有值"] = old_value
                if old_value != new_value:
                    changes.append({"field": "含有值" if sheet_name != "包材" else "允許濃度", "old": old_value, "new": new_value})
                for field in COMPARE_FIELDS:
                    if values_differ(field, old.get(field), new.get(field)):
                        changes.append({"field": field, "old": normalize_compare_value(old.get(field)), "new": normalize_compare_value(new.get(field))})
                diff_type = "modified" if changes else "same"
            else:
                old = {}
                diff_type = "added"
        summary[diff_type] += 1
        composite_key = make_key(new, [*SUBSTANCE_KEY_FIELDS, "供應商代號", "物料名稱"])
        results.append({"diff_type": diff_type, "composite_key": composite_key, "old": old, "new": new, "changes": changes})
    wb.close()
    return results, summary


def duplicate_keys(records: list[dict], fields: list[str]) -> set[str]:
    counts: dict[str, int] = {}
    for r in records:
        key = make_key(r, fields)
        counts[key] = counts.get(key, 0) + 1
    return {k for k, v in counts.items() if v > 1}


def compare_records(old_records: list[dict], new_records: list[dict], key_fields: list[str]) -> tuple[list[dict], dict]:
    old_dupes = duplicate_keys(old_records, key_fields)
    new_dupes = duplicate_keys(new_records, key_fields)
    old_map = {make_key(r, key_fields): r for r in old_records if make_key(r, key_fields) not in old_dupes}
    new_map = {make_key(r, key_fields): r for r in new_records if make_key(r, key_fields) not in new_dupes}
    fallback_fields = [field for field in key_fields if field != "CAS主號"]
    unmatched_old_keys = set(old_map) - set(new_map)
    unmatched_new_keys = set(new_map) - set(old_map)
    old_fallback: dict[str, list[str]] = {}
    new_fallback: dict[str, list[str]] = {}
    for key in unmatched_old_keys:
        fallback_key = make_key(old_map[key], fallback_fields)
        old_fallback.setdefault(fallback_key, []).append(key)
    for key in unmatched_new_keys:
        fallback_key = make_key(new_map[key], fallback_fields)
        new_fallback.setdefault(fallback_key, []).append(key)
    fallback_pairs: dict[str, str] = {}
    fallback_new_keys: set[str] = set()
    for fallback_key, old_keys in old_fallback.items():
        new_keys = new_fallback.get(fallback_key, [])
        if len(old_keys) == 1 and len(new_keys) == 1:
            fallback_pairs[old_keys[0]] = new_keys[0]
            fallback_new_keys.add(new_keys[0])
    remaining_old_keys = unmatched_old_keys - set(fallback_pairs)
    remaining_new_keys = unmatched_new_keys - fallback_new_keys
    cas_fields = [field for field in key_fields if field != "化學物質"]
    old_cas_fallback: dict[str, list[str]] = {}
    new_cas_fallback: dict[str, list[str]] = {}
    for key in remaining_old_keys:
        if normalize_key_value("CAS主號", old_map[key].get("CAS主號")) in {"", "-", "--", "_", "同上"}:
            continue
        cas_key = make_key(old_map[key], cas_fields)
        old_cas_fallback.setdefault(cas_key, []).append(key)
    for key in remaining_new_keys:
        if normalize_key_value("CAS主號", new_map[key].get("CAS主號")) in {"", "-", "--", "_", "同上"}:
            continue
        cas_key = make_key(new_map[key], cas_fields)
        new_cas_fallback.setdefault(cas_key, []).append(key)
    for cas_key, old_keys in old_cas_fallback.items():
        new_keys = new_cas_fallback.get(cas_key, [])
        if len(old_keys) == 1 and len(new_keys) == 1:
            fallback_pairs[old_keys[0]] = new_keys[0]
            fallback_new_keys.add(new_keys[0])
    results = []
    all_keys = sorted(((set(new_map) - fallback_new_keys) | set(fallback_pairs) | new_dupes))
    summary = {"added": 0, "removed": 0, "modified": 0, "same": 0, "conflict": 0}
    for key in all_keys:
        old = old_map.get(key, {})
        new_key = fallback_pairs.get(key, key)
        new = new_map.get(new_key, {})
        changes = []
        if new.get("資料問題"):
            diff_type = "conflict"
            changes.append({"field": "資料問題", "old": "", "new": new.get("資料問題")})
        elif old.get("公式儲存格"):
            diff_type = "conflict"
            changes.append({"field": "公式儲存格", "old": old.get("含有值", ""), "new": "目標儲存格為公式，需人工確認"})
        elif key in old_dupes or key in new_dupes:
            diff_type = "conflict"
            changes.append({"field": "Composite Key", "old": "重複" if key in old_dupes else "", "new": "重複" if key in new_dupes else ""})
        elif new and not old:
            diff_type = "added"
        else:
            for field in COMPARE_FIELDS:
                if values_differ(field, old.get(field), new.get(field)):
                    changes.append({"field": field, "old": normalize_compare_value(old.get(field)), "new": normalize_compare_value(new.get(field))})
            diff_type = "modified" if changes else "same"
        summary[diff_type] += 1
        composite_key = key if new_key == key else f"{key} -> {new_key}"
        results.append({"diff_type": diff_type, "composite_key": composite_key, "old": old, "new": new, "changes": changes})
    return results, summary


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    if source_row in ws.row_dimensions:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)


def material_value_columns(ws, sheet_name: str) -> list[int]:
    if sheet_name == "原物料":
        return [col for col in range(6, ws.max_column + 1) if clean_text(ws.cell(5, col).value)]
    layout = package_layout(ws)
    return [col for col in range(layout["material_start"], ws.max_column + 1) if clean_text(ws.cell(layout["material_row"], col).value)]


def supplier_code_for_column(ws, sheet_name: str, col: int) -> str:
    if sheet_name == "包材":
        return package_supplier_code(ws, col)
    supplier_row = 7 if sheet_name == "原物料" else 3
    return clean_text(ws.cell(supplier_row, col).value)


def supplier_name_for_column(ws, sheet_name: str, col: int) -> str:
    if sheet_name == "包材":
        return package_supplier_name(ws, col)
    return clean_text(ws.cell(8, col).value)


def supplier_identifier_matches_column(ws, sheet_name: str, col: int, supplier_identifier: str) -> bool:
    supplier_identifier = clean_text(supplier_identifier)
    if not supplier_identifier:
        return False
    if sheet_name == "包材":
        return package_supplier_matches(ws, col, supplier_identifier)
    return normalize_key_value("供應商代號", supplier_code_for_column(ws, sheet_name, col)) == normalize_key_value("供應商代號", supplier_identifier)


def fill_new_substance_supplier_values(
    ws,
    sheet_name: str,
    row_num: int,
    supplier_code: str,
    target_col: int | None,
    target_value: str,
) -> int:
    updated = 0
    for col in material_value_columns(ws, sheet_name):
        cell = ws.cell(row_num, col)
        if col == target_col or supplier_identifier_matches_column(ws, sheet_name, col, supplier_code):
            if target_value:
                cell.value = target_value
                updated += 1
            else:
                cell.value = "no contain"
        else:
            cell.value = "no contain"
    return updated


def copy_column_style(ws, source_col: int, target_col: int) -> None:
    source_letter = excel_col_name(source_col)
    target_letter = excel_col_name(target_col)
    if source_letter in ws.column_dimensions:
        ws.column_dimensions[target_letter].width = ws.column_dimensions[source_letter].width
    for row in range(1, ws.max_row + 1):
        src = ws.cell(row, source_col)
        dst = ws.cell(row, target_col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)


def create_material_column(ws, sheet_name: str, material: str, supplier_code: str, supplier: str, category: str) -> int:
    col = ws.max_column + 1
    copy_column_style(ws, col - 1, col)
    if sheet_name == "原物料":
        header_values = {5: material, 6: category, 7: supplier_code, 8: supplier}
        data_start, substance_col, cas_col = 9, 3, 4
    else:
        layout = package_layout(ws)
        supplier_value = supplier or supplier_code
        header_values = {layout["material_row"]: material}
        if layout.get("supplier_code_row") and looks_like_supplier_code(supplier_code):
            header_values[layout["supplier_code_row"]] = supplier_code
        if layout.get("supplier_row"):
            header_values[layout["supplier_row"]] = supplier_value
        data_start, substance_col, cas_col = layout["data_start"], 2, 3
    for row, value in header_values.items():
        ws.cell(row, col).value = value
    for row in range(data_start, ws.max_row + 1):
        if clean_text(ws.cell(row, substance_col).value) or clean_text(ws.cell(row, cas_col).value):
            ws.cell(row, col).value = "no contain"
    return col


def build_report(job_id: str, output: Path) -> None:
    with connect() as conn:
        job = conn.execute("select * from comparison_jobs where id=?", (job_id,)).fetchone()
        rows = conn.execute("select * from comparison_results where job_id=?", (job_id,)).fetchall()
    summary = json.loads(job["summary_json"])
    grouped = {"added": [], "removed": [], "modified": [], "conflict": [], "same": []}
    modified_rows = []
    full = []
    for row in rows:
        old = json.loads(row["old_json"])
        new = json.loads(row["new_json"])
        changes = json.loads(row["changes_json"])
        base = {
            "差異類型": row["diff_type"],
            "Composite Key": row["composite_key"],
            "處理方式": row["action"],
            "廠商代號": new.get("供應商代號") or old.get("供應商代號"),
            "物料名稱": new.get("物料名稱") or old.get("物料名稱"),
            "化學物質": new.get("化學物質") or old.get("化學物質"),
            "CAS": new.get("CAS") or old.get("CAS"),
            "舊值": old.get("含有值", ""),
            "新值": new.get("含有值", ""),
        }
        grouped[row["diff_type"]].append(base)
        full.append(base)
        for change in changes:
            modified_rows.append({**base, "欄位名稱": change["field"], "修改前": change["old"], "修改後": change["new"]})
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(
            [
                {"項目": "比對時間", "值": job["created_at"]},
                {"項目": "舊版總筆數", "值": summary.get("old_count")},
                {"項目": "新版總筆數", "值": summary.get("new_count")},
                {"項目": "新增筆數", "值": summary.get("added")},
                {"項目": "刪除筆數", "值": summary.get("removed")},
                {"項目": "修改筆數", "值": summary.get("modified")},
                {"項目": "相同筆數", "值": summary.get("same")},
                {"項目": "衝突筆數", "值": summary.get("conflict")},
            ]
        ).to_excel(writer, sheet_name="Summary", index=False)
        pd.DataFrame(grouped["added"]).to_excel(writer, sheet_name="Added", index=False)
        pd.DataFrame(grouped["removed"]).to_excel(writer, sheet_name="Inactive", index=False)
        pd.DataFrame(modified_rows).to_excel(writer, sheet_name="Modified", index=False)
        pd.DataFrame(grouped["conflict"]).to_excel(writer, sheet_name="Conflicts", index=False)
        pd.DataFrame(full).to_excel(writer, sheet_name="Full Comparison", index=False)


def find_material_column(ws, sheet_name: str, material: str, supplier_code: str) -> int | None:
    if sheet_name == "原物料":
        for col in range(6, ws.max_column + 1):
            if normalize_key_value("物料名稱", ws.cell(5, col).value) == normalize_key_value("物料名稱", material):
                if not supplier_code or normalize_key_value("供應商代號", ws.cell(7, col).value) == normalize_key_value("供應商代號", supplier_code):
                    return col
    if sheet_name == "包材":
        layout = package_layout(ws)
        for col in range(layout["material_start"], ws.max_column + 1):
            if normalize_key_value("物料名稱", ws.cell(layout["material_row"], col).value) == normalize_key_value("物料名稱", material):
                if package_supplier_matches(ws, col, supplier_code, layout):
                    return col
    return None


def apply_job(job_id: str) -> dict:
    with connect() as conn:
        job = conn.execute("select * from comparison_jobs where id=?", (job_id,)).fetchone()
        master = conn.execute("select * from excel_files where id=?", (job["master_file_id"],)).fetchone()
        rows = conn.execute("select * from comparison_results where job_id=? and action != 'pending'", (job_id,)).fetchall()
        all_rows = conn.execute("select * from comparison_results where job_id=?", (job_id,)).fetchall()
    source = Path(master["stored_path"])
    version = next_master_version()
    output = OUTPUTS / f"物料總表_{version}.xlsx"
    shutil.copy2(source, output)
    wb = openpyxl.load_workbook(output)
    sheet_name = job["master_sheet"]
    sheet_name = resolve_sheet_name(wb, sheet_name)
    ws = wb[sheet_name]
    summary = json.loads(job["summary_json"])
    if summary.get("mode") == "substance_rows":
        existing = parse_master_substances(output, sheet_name)
        row_by_key = {make_key(r, SUBSTANCE_KEY_FIELDS): r["excel_row"] for r in existing}
        changed = {"added": 0, "modified": 0, "removed": 0, "columns_added": 0, "content_updated": 0}
        newly_added_row_keys: set[str] = set()
        batch_mode = bool(summary.get("batch"))
        material = clean_text(summary.get("material_name") or job["material_name"])
        supplier_code = clean_text(summary.get("supplier_code") or job["supplier_code"])
        supplier = clean_text(summary.get("supplier"))
        horizontal_col = None
        if material and supplier_code:
            horizontal_col = find_material_column(ws, sheet_name, material, supplier_code)
            if horizontal_col is None and summary.get("horizontal_new_column"):
                horizontal_col = create_material_column(ws, sheet_name, material, supplier_code, supplier, job["category"])
                changed["columns_added"] += 1
        for result in rows:
            diff_type = result["diff_type"]
            action = result["action"]
            old = json.loads(result["old_json"])
            new = json.loads(result["new_json"])
            record = new or old
            if diff_type == "conflict" or action in {"keep_old", "ignore"}:
                continue
            row_key = make_key(enrich_substance_key(record), SUBSTANCE_KEY_FIELDS)
            row_num = row_by_key.get(row_key) or old.get("excel_row")
            if action in {"accept_new", "confirm_add"} and new:
                target_col = horizontal_col
                target_supplier_code = supplier_code
                target_value = effective_survey_update_value(new, sheet_name)
                if batch_mode:
                    target_material = clean_text(new.get("物料名稱"))
                    target_supplier_code = clean_text(new.get("供應商代號"))
                    target_supplier = clean_text(new.get("供應商"))
                    target_col = find_material_column(ws, sheet_name, target_material, target_supplier_code)
                    if target_col is None and target_material and target_supplier_code:
                        target_col = create_material_column(ws, sheet_name, target_material, target_supplier_code, target_supplier, job["category"])
                        changed["columns_added"] += 1
                is_new_row = not row_num
                if not row_num:
                    row_num = ws.max_row + 1
                    copy_row_style(ws, row_num - 1, row_num)
                    row_by_key[row_key] = row_num
                    newly_added_row_keys.add(row_key)
                    changed["added"] += 1
                elif diff_type == "modified":
                    changed["modified"] += 1
                if sheet_name == "原物料":
                    values = {2: new.get("項次"), 3: new.get("化學物質"), 4: new.get("CAS"), 5: new.get("等級")}
                else:
                    values = {1: new.get("項次"), 2: new.get("化學物質"), 3: new.get("CAS"), 4: new.get("等級")}
                for c, value in values.items():
                    if value != "":
                        ws.cell(row_num, c).value = value
                if is_new_row:
                    changed["content_updated"] += fill_new_substance_supplier_values(
                        ws,
                        sheet_name,
                        row_num,
                        target_supplier_code,
                        target_col,
                        target_value,
                    )
                elif batch_mode and target_col and target_value:
                    ws.cell(row_num, target_col).value = target_value
                    changed["content_updated"] += 1
                row_by_key[make_key(enrich_substance_key(new), SUBSTANCE_KEY_FIELDS)] = row_num
            elif action in {"mark_inactive", "confirm_remove"}:
                changed["removed"] += 1
        if horizontal_col and not batch_mode:
            for result in all_rows:
                if result["diff_type"] == "conflict":
                    continue
                action = result["action"]
                if action in {"keep_old", "ignore", "confirm_remove", "mark_inactive"}:
                    continue
                new = json.loads(result["new_json"])
                if not new:
                    continue
                row_key = make_key(enrich_substance_key(new), SUBSTANCE_KEY_FIELDS)
                if row_key in newly_added_row_keys:
                    continue
                row_num = row_by_key.get(row_key) or new.get("excel_row")
                value = effective_survey_update_value(new, sheet_name)
                if row_num and value:
                    ws.cell(row_num, horizontal_col).value = value
                    changed["content_updated"] += 1
        wb.save(output)
        wb.close()
        report = OUTPUTS / f"Search_Report_{datetime.now().strftime('%Y%m%d')}_{job_id[:8]}.xlsx"
        build_report(job_id, report)
        file_id = str(uuid.uuid4())
        with connect() as conn:
            conn.execute(
                "insert into excel_files values (?, ?, ?, ?, ?, ?, ?)",
                (file_id, "updated_master", output.name, str(output), file_hash(output), now(), json.dumps({"report": str(report)}, ensure_ascii=False)),
            )
            conn.execute(
                "insert into excel_versions values (?, ?, ?, ?, ?, ?, ?)",
                (str(uuid.uuid4()), version, file_id, None, now(), "管理者", f"套用細項列比對工作 {job_id}"),
            )
            set_state(conn, "working_master_file_id", file_id)
        log("套用更新", f"產生細項列更新版 {output.name}", job_id)
        return {"updated_file": output.name, "updated_file_id": file_id, "report_file": report.name, "changed": changed}
    sample_new = next((json.loads(r["new_json"]) for r in rows if json.loads(r["new_json"])), {})
    material = sample_new.get("物料名稱") or job["material_name"]
    supplier_code = sample_new.get("供應商代號") or job["supplier_code"]
    col = int(summary.get("target_column") or 0)
    if not col:
        col = find_material_column(ws, sheet_name, material, supplier_code) or 0
    if col is None:
        col = ws.max_column + 1
        if sheet_name == "原物料":
            for src_row, value in [(5, material), (6, job["category"]), (7, supplier_code), (8, sample_new.get("供應商", ""))]:
                ws.cell(src_row, col).value = value
                ws.cell(src_row, col)._style = copy(ws.cell(src_row, col - 1)._style)
        else:
            for src_row, value in [(3, supplier_code), (4, sample_new.get("供應商", "")), (5, material)]:
                ws.cell(src_row, col).value = value
                ws.cell(src_row, col)._style = copy(ws.cell(src_row, col - 1)._style)
    existing = parse_master_column(output, sheet_name, col)
    row_by_key = {make_key(r, SUBSTANCE_KEY_FIELDS): r["excel_row"] for r in existing}
    changed = {"added": 0, "modified": 0, "removed": 0}
    for result in rows:
        diff_type = result["diff_type"]
        action = result["action"]
        old = json.loads(result["old_json"])
        new = json.loads(result["new_json"])
        record = new or old
        if diff_type == "conflict" or action in {"keep_old", "ignore"}:
            continue
        row_key = make_key(enrich_substance_key(record), SUBSTANCE_KEY_FIELDS)
        row_num = row_by_key.get(row_key)
        if row_num is None and action in {"accept_new", "confirm_add"}:
            row_num = ws.max_row + 1
            copy_row_style(ws, row_num - 1, row_num)
            if sheet_name == "原物料":
                values = {2: record.get("項次"), 3: record.get("化學物質"), 4: record.get("CAS"), 5: record.get("等級")}
            else:
                values = {1: record.get("項次"), 2: record.get("化學物質"), 3: record.get("CAS"), 4: record.get("等級")}
            for c, value in values.items():
                ws.cell(row_num, c).value = value
            fill_new_substance_supplier_values(
                ws,
                sheet_name,
                row_num,
                supplier_code,
                col,
                normalize_compare_value(record.get("含有值")),
            )
            row_by_key[row_key] = row_num
            changed["added"] += 1
        if row_num:
            if action in {"accept_new", "confirm_add"}:
                new_value = new.get("含有值", "")
                if new_value:
                    ws.cell(row_num, col).value = new_value
                if diff_type == "modified":
                    changed["modified"] += 1
            elif action in {"mark_inactive", "confirm_remove"}:
                ws.cell(row_num, col).value = "停用"
                changed["removed"] += 1
    wb.save(output)
    wb.close()
    report = OUTPUTS / f"Search_Report_{supplier_code or 'UNKNOWN'}_{datetime.now().strftime('%Y%m%d')}_{job_id[:8]}.xlsx"
    build_report(job_id, report)
    file_id = str(uuid.uuid4())
    with connect() as conn:
        conn.execute(
            "insert into excel_files values (?, ?, ?, ?, ?, ?, ?)",
            (file_id, "updated_master", output.name, str(output), file_hash(output), now(), json.dumps({"report": str(report)}, ensure_ascii=False)),
        )
        conn.execute(
            "insert into excel_versions values (?, ?, ?, ?, ?, ?, ?)",
            (str(uuid.uuid4()), version, file_id, None, now(), "管理者", f"套用比對工作 {job_id}"),
        )
        set_state(conn, "working_master_file_id", file_id)
    log("建立版本", f"已套用本批修改並產生 {output.name}", job_id)
    return {"updated_file": output.name, "updated_file_id": file_id, "report_file": report.name, "changed": changed}


def json_response(handler: BaseHTTPRequestHandler, data, status=200) -> None:
    body = json.dumps(data, ensure_ascii=False, default=str).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def error_response(handler: BaseHTTPRequestHandler, message: str, status=400) -> None:
    json_response(handler, {"error": message}, status)


class App(BaseHTTPRequestHandler):
    def do_GET(self):
        try:
            parsed = urlparse(self.path)
            path = unquote(parsed.path)
            if path == "/api/dashboard":
                return self.dashboard()
            if path.startswith("/api/files/"):
                return self.file_detail(path.split("/")[-1])
            if path == "/api/key-settings":
                return self.get_key_settings()
            if path.startswith("/api/jobs/") and path.endswith("/results"):
                return self.job_results(path.split("/")[-2], parse_qs(parsed.query))
            if path.startswith("/download/"):
                return self.download(path.removeprefix("/download/"))
            if path.startswith("/download-file/"):
                return self.download_file(path.removeprefix("/download-file/"))
            return self.static(path)
        except Exception as exc:
            traceback.print_exc()
            error_response(self, f"系統處理失敗：{exc}", 500)

    def do_POST(self):
        try:
            path = urlparse(self.path).path
            if path == "/api/upload":
                return self.upload()
            if path == "/api/compare":
                return self.compare()
            if path == "/api/material-candidates":
                return self.get_material_candidates()
            if path == "/api/material-columns":
                return self.get_material_columns()
            if path == "/api/master-search":
                return self.master_search()
            if path == "/api/master-filter-options":
                return self.master_filter_options()
            if path == "/api/key-settings":
                return self.save_key_settings()
            if path.startswith("/api/files/") and path.endswith("/set-latest"):
                return self.set_latest_master(path.split("/")[-2])
            if path == "/api/finalize-working-version":
                return self.finalize_working_version()
            if path == "/api/clear-working-version":
                return self.clear_working_version()
            if path.startswith("/api/jobs/") and path.endswith("/confirm"):
                return self.confirm(path.split("/")[-2])
            if path.startswith("/api/jobs/") and path.endswith("/auto-confirm"):
                return self.auto_confirm(path.split("/")[-2])
            if path.startswith("/api/jobs/") and path.endswith("/cancel"):
                return self.cancel_job(path.split("/")[-2])
            if path.startswith("/api/jobs/") and path.endswith("/apply"):
                return self.apply(path.split("/")[-2])
            error_response(self, "找不到此功能", 404)
        except Exception as exc:
            traceback.print_exc()
            error_response(self, f"系統處理失敗：{exc}", 500)

    def do_DELETE(self):
        try:
            path = urlparse(self.path).path
            if path.startswith("/api/files/"):
                return self.delete_file(path.split("/")[-1])
            error_response(self, "找不到此功能", 404)
        except Exception as exc:
            traceback.print_exc()
            error_response(self, f"系統處理失敗：{exc}", 500)

    def do_Delete(self):
        return self.do_DELETE()

    def parse_body(self) -> dict:
        length = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(length)
        return json.loads(raw.decode("utf-8")) if raw else {}

    def dashboard(self):
        with connect() as conn:
            state = version_state(conn)
            pinned_ids = [item["id"] for item in [state.get("latest_master"), state.get("working_master")] if item]
            files_by_id = {}
            if pinned_ids:
                placeholders = ",".join("?" for _ in pinned_ids)
                for row in conn.execute(f"select * from excel_files where id in ({placeholders})", pinned_ids):
                    files_by_id[row["id"]] = dict(row)
            for row in conn.execute("select * from excel_files order by uploaded_at desc limit 10"):
                files_by_id[row["id"]] = dict(row)
            files = list(files_by_id.values())
            jobs = [dict(r) for r in conn.execute("select * from comparison_jobs order by created_at desc limit 10")]
            versions = [dict(r) for r in conn.execute("select * from excel_versions order by created_at desc limit 10")]
        return json_response(
            self,
            {
                "files": files,
                "jobs": jobs,
                "versions": versions,
                "version_state": state,
                "system_notices": state.get("notices", []),
                "app_version": APP_VERSION,
                "app_version_note": APP_VERSION_NOTE,
            },
        )

    def set_latest_master(self, file_id: str):
        with connect() as conn:
            row = conn.execute("select * from excel_files where id=?", (file_id,)).fetchone()
            if not row or row["kind"] not in {"master", "updated_master"}:
                return error_response(self, "只能將總表或更動版總表設為最新版本")
            latest_version = conn.execute(
                "select * from excel_versions order by created_at desc limit 1"
            ).fetchone()
            candidate_version = conn.execute(
                "select * from excel_versions where file_id=? order by created_at desc limit 1",
                (file_id,),
            ).fetchone()
            if latest_version and candidate_version and candidate_version["created_at"] < latest_version["created_at"]:
                return error_response(
                    self,
                    "此檔案早於目前版本歷程中的最新定版，不能直接設回最新版；若確定要回復舊版，請先重新上傳該總表作為新版本。",
                )
            set_state(conn, "latest_master_file_id", file_id)
            set_state(conn, "working_master_file_id", "")
            conn.execute(
                "update excel_versions set note=case when instr(coalesce(note, ''), '已定為最新版本')=0 then coalesce(note, '') || '；已定為最新版本' else note end where file_id=?",
                (file_id,),
            )
        log("設定最新版本", f"已將 {row['original_name']} 設為最新總表", file_id)
        return json_response(self, {"ok": True, "latest_master": dict(row), "message": f"已設定為目前最新總表：{row['original_name']}"})

    def finalize_working_version(self):
        with connect() as conn:
            state = version_state(conn)
            working = state.get("working_master")
            if not working:
                return error_response(self, "尚未有本輪暫存更動版，請先完成至少一次比對並套用更新。")
            file_id = working["id"]
            set_state(conn, "latest_master_file_id", file_id)
            set_state(conn, "working_master_file_id", "")
            conn.execute(
                "update excel_versions set note=case when instr(coalesce(note, ''), '已定為最新版本')=0 then coalesce(note, '') || '；已定為最新版本' else note end where file_id=?",
                (file_id,),
            )
        log("完成本輪更新並定版", f"已將 {working['original_name']} 設為最新總表", file_id)
        return json_response(self, {"ok": True, "latest_master": working, "message": f"已完成本輪更新並定版：{working['original_name']}"})

    def clear_working_version(self):
        stored_path = None
        with connect() as conn:
            state = version_state(conn)
            working = state.get("working_master")
            if working:
                row = conn.execute("select * from excel_files where id=?", (working["id"],)).fetchone()
                if row:
                    stored_path = Path(row["stored_path"])
                    conn.execute("delete from excel_versions where file_id=?", (working["id"],))
                    conn.execute("delete from excel_files where id=?", (working["id"],))
            set_state(conn, "working_master_file_id", "")
        if working:
            if stored_path:
                stored_path.unlink(missing_ok=True)
            log("改選比對基準", f"清除本輪暫存基準 {working['original_name']}", working["id"])
        return json_response(self, {"ok": True})

    def get_key_settings(self):
        with connect() as conn:
            row = conn.execute("select fields_json from composite_key_settings where id=1").fetchone()
        allowed_fields = {"來源", "化學物質", "CAS主號", "等級"}
        fields = [field for field in json.loads(row["fields_json"]) if field in allowed_fields]
        if "等級" not in fields:
            fields.append("等級")
        return json_response(self, {"fields": fields})

    def save_key_settings(self):
        body = self.parse_body()
        allowed_fields = {"來源", "化學物質", "CAS主號", "等級"}
        fields = [field for field in (body.get("fields") or DEFAULT_KEY_FIELDS) if field in allowed_fields] or DEFAULT_KEY_FIELDS
        if "等級" not in fields:
            fields.append("等級")
        with connect() as conn:
            conn.execute("update composite_key_settings set fields_json=? where id=1", (json.dumps(fields, ensure_ascii=False),))
        log("修改比對欄位設定", "、".join(fields))
        return json_response(self, {"ok": True, "fields": fields})

    def upload(self):
        ctype, pdict = cgi.parse_header(self.headers.get("Content-Type"))
        if ctype != "multipart/form-data":
            return error_response(self, "請使用表單上傳 Excel 檔案")
        pdict["boundary"] = bytes(pdict["boundary"], "utf-8")
        form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": self.headers["Content-Type"]})
        file_item = form["file"]
        kind = clean_text(form.getvalue("kind") or "master")
        if not file_item.filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
            return error_response(self, "檔案格式不支援，請上傳 .xlsx、.xlsm 或 .xls")
        file_id = str(uuid.uuid4())
        original = Path(file_item.filename).name
        stored = UPLOADS / f"{file_id}_{original}"
        with stored.open("wb") as f:
            shutil.copyfileobj(file_item.file, f)
        try:
            info = workbook_info(stored)
        except Exception as exc:
            stored.unlink(missing_ok=True)
            return error_response(self, f"Excel 無法開啟，請確認檔案未加密且未損壞：{exc}")
        with connect() as conn:
            conn.execute(
                "insert into excel_files values (?, ?, ?, ?, ?, ?, ?)",
                (file_id, kind, original, str(stored), file_hash(stored), now(), json.dumps(info, ensure_ascii=False)),
            )
            if kind == "master":
                version = "MASTER-" + datetime.now().strftime("%Y%m%d") + "-001"
                conn.execute("insert into excel_versions values (?, ?, ?, ?, ?, ?, ?)", (str(uuid.uuid4()), version, file_id, None, now(), "管理者", "上傳總表"))
        log("上傳", f"上傳 {original}", file_id)
        return json_response(self, {"id": file_id, "name": original, "kind": kind, "info": info})

    def file_detail(self, file_id: str):
        with connect() as conn:
            row = conn.execute("select * from excel_files where id=?", (file_id,)).fetchone()
        if not row:
            return error_response(self, "找不到檔案", 404)
        return json_response(self, {**dict(row), "metadata": json.loads(row["metadata_json"])})

    def delete_file(self, file_id: str):
        with connect() as conn:
            row = conn.execute("select * from excel_files where id=?", (file_id,)).fetchone()
            if not row:
                return error_response(self, "找不到檔案", 404)
            used_count = conn.execute(
                """
                select count(*) c
                from comparison_jobs
                where (master_file_id=? or survey_file_id=?)
                  and status in ('processing', 'applying')
                """,
                (file_id, file_id),
            ).fetchone()["c"]
            if used_count:
                return error_response(self, "此檔案正在比對或套用中，請等待工作完成後再刪除。")
            latest_id = get_state(conn, "latest_master_file_id")
            working_id = get_state(conn, "working_master_file_id")
            if file_id == latest_id:
                return error_response(self, "目前最新總表不能直接刪除；請先將其他總表定為最新版。")
            if file_id == working_id:
                return error_response(self, "本輪暫存更動版請用取消本次比對清除。")
            if row["kind"] == "updated_master":
                return error_response(self, "已產生的更動版總表會保留在版本歷程中，不能用一般刪除移除。")
            stored_path = Path(row["stored_path"])
            conn.execute("delete from excel_versions where file_id=?", (file_id,))
            conn.execute(
                "update comparison_jobs set status='file_deleted' where (master_file_id=? or survey_file_id=?) and status!='applied'",
                (file_id, file_id),
            )
            conn.execute("delete from excel_files where id=?", (file_id,))
        try:
            stored_path.unlink(missing_ok=True)
        except OSError as exc:
            return error_response(self, f"資料庫紀錄已移除，但檔案刪除失敗：{exc}", 500)
        log("刪除上傳檔案", f"刪除未使用檔案 {row['original_name']}", file_id)
        return json_response(self, {"ok": True, "deleted_id": file_id})

    def compare(self):
        body = self.parse_body()
        master_id, survey_id = body.get("master_file_id"), body.get("survey_file_id")
        survey_items = body.get("survey_items") or []
        with connect() as conn:
            master = conn.execute("select * from excel_files where id=?", (master_id,)).fetchone()
            survey = conn.execute("select * from excel_files where id=?", (survey_id,)).fetchone() if survey_id else None
            survey_files = {
                row["id"]: row
                for row in conn.execute(
                    f"select * from excel_files where id in ({','.join('?' for _ in survey_items)})",
                    [item.get("survey_file_id") for item in survey_items],
                ).fetchall()
            } if survey_items else {}
        if not master or (not survey and not survey_items):
            return error_response(self, "請先上傳總表與調查表")
        master_sheet = body.get("master_sheet") or "原物料"
        workbook_for_column = openpyxl.load_workbook(Path(master["stored_path"]), read_only=True, data_only=False)
        resolved_sheet = resolve_sheet_name(workbook_for_column, master_sheet)
        category = "包材" if resolved_sheet == "包材" else ""
        subcategory = ""
        survey_version = clean_text(body.get("survey_version"))
        survey_source = "包材" if resolved_sheet == "包材" else "原物料"
        warnings = []
        if survey_items:
            new_records = []
            survey_meta = {"supplier": "", "warnings": [], "rows": 0}
            for index, item in enumerate(survey_items, start=1):
                file_id = item.get("survey_file_id")
                file_row = survey_files.get(file_id)
                if not file_row:
                    warnings.append(f"第 {index} 筆調查表檔案不存在，已略過")
                    continue
                supplier_code = clean_text(item.get("supplier_code"))
                material_name = clean_text(item.get("material_name"))
                sub_name = clean_text(item.get("subcategory"))
                if not supplier_code:
                    workbook_for_column.close()
                    field_name = "供應商名稱/代號" if resolved_sheet == "包材" else "供應商代號"
                    return error_response(self, f"{file_row['original_name']} 未填{field_name}")
                if not material_name:
                    workbook_for_column.close()
                    return error_response(self, f"{file_row['original_name']} 未填品項/原料名稱")
                parsed_records, meta = parse_survey(Path(file_row["stored_path"]), supplier_code, survey_source, sub_name, survey_version)
                item_supplier = clean_text(meta.get("supplier")) or supplier_code
                for record in parsed_records:
                    record["來源"] = resolved_sheet
                    record["物料名稱"] = material_name
                    record["物料編號"] = material_name
                    record["細項類別"] = sub_name
                    record["來源檔案"] = file_row["original_name"]
                    record["調查表類型"] = resolved_sheet
                    if resolved_sheet == "包材":
                        record["供應商"] = item_supplier
                        record["含有值"] = normalize_compare_value(record.get("允許濃度"))
                    new_records.append(enrich_substance_key(record))
                warnings.extend(meta.get("warnings", []))
            old_records = parse_master_substances(Path(master["stored_path"]), resolved_sheet)
            results, summary = compare_batch_records(Path(master["stored_path"]), resolved_sheet, new_records, old_records)
            material_name = ""
            supplier_code = ""
        else:
            supplier_code = clean_text(body.get("supplier_code"))
            new_records, survey_meta = parse_survey(Path(survey["stored_path"]), supplier_code, survey_source, subcategory, survey_version)
            old_records = parse_master_substances(Path(master["stored_path"]), resolved_sheet)
            material_name = clean_text(body.get("material_name")) or survey_meta.get("material_name", "")
            row_key_fields = [field for field in DEFAULT_KEY_FIELDS if field in {"來源", "化學物質", "CAS主號", "等級"}]
            actual_sheet = old_records[0]["來源"] if old_records else master_sheet
            for record in new_records:
                record["來源"] = actual_sheet
                if material_name:
                    record["物料名稱"] = material_name
                    record["物料編號"] = material_name
                if not record.get("化學物質類別"):
                    record["化學物質類別"] = category
                if not record.get("細項類別"):
                    record["細項類別"] = subcategory
            results, summary = compare_records(old_records, new_records, row_key_fields)
            warnings = list(survey_meta.get("warnings", []))
        allowed_key_fields = {"來源", "化學物質", "CAS主號", "等級"}
        row_key_fields = [field for field in DEFAULT_KEY_FIELDS if field in allowed_key_fields]
        actual_sheet = resolved_sheet
        horizontal_column = None
        horizontal_new_column = False
        horizontal_warning = ""
        if not survey_items and material_name and supplier_code and resolved_sheet in workbook_for_column.sheetnames:
            horizontal_column = find_material_column(workbook_for_column[resolved_sheet], resolved_sheet, material_name, supplier_code)
            horizontal_new_column = horizontal_column is None
        elif not survey_items and (material_name or supplier_code):
            supplier_field = "供應商名稱/代號" if resolved_sheet == "包材" else "供應商代號"
            horizontal_warning = f"未同時取得品名與{supplier_field}，本次不會自動新增橫向供應品項欄位。"
        if horizontal_warning:
            warnings.append(horizontal_warning)
        workbook_for_column.close()
        summary.update(
            {
                "old_count": len(old_records),
                "new_count": len(new_records),
                "warnings": warnings,
                "mode": "substance_rows",
                "key_fields": row_key_fields,
                "material_name": material_name,
                "supplier_code": supplier_code,
                "supplier": survey_meta.get("supplier", ""),
                "horizontal_column": horizontal_column,
                "horizontal_new_column": horizontal_new_column,
                "batch": bool(survey_items),
            }
        )
        job_id = str(uuid.uuid4())
        with connect() as conn:
            conn.execute(
                "insert into comparison_jobs values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (job_id, master_id, survey_id or (survey_items[0].get("survey_file_id") if survey_items else ""), actual_sheet, material_name, supplier_code, category, subcategory, "pending_confirmation", now(), json.dumps(summary, ensure_ascii=False)),
            )
            for result in results:
                conn.execute(
                    "insert into comparison_results values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (
                        str(uuid.uuid4()),
                        job_id,
                        result["diff_type"],
                        result["composite_key"],
                        json.dumps(result["old"], ensure_ascii=False),
                        json.dumps(result["new"], ensure_ascii=False),
                        json.dumps(result["changes"], ensure_ascii=False),
                        "pending",
                        "",
                        None,
                    ),
                )
        log("比對", f"建立比對工作 {job_id}", job_id)
        return json_response(self, {"job_id": job_id, "summary": summary, "survey_meta": survey_meta})

    def master_search(self):
        body = self.parse_body()
        master_id = body.get("master_file_id")
        with connect() as conn:
            master = conn.execute("select * from excel_files where id=?", (master_id,)).fetchone()
        if not master:
            return error_response(self, "請先選用要查詢的總表")
        sheet_name = body.get("master_sheet") or "原物料"
        filters = {
            "supplier_code": body.get("supplier_code"),
            "supplier": body.get("supplier"),
            "material_name": body.get("material_name"),
            "category": body.get("category"),
            "substance": body.get("substance"),
            "include_empty": body.get("include_empty"),
            "page": body.get("page"),
            "page_size": body.get("page_size"),
        }
        result = query_master_details(Path(master["stored_path"]), sheet_name, filters)
        result["file_name"] = master["original_name"]
        result["file_kind"] = master["kind"]
        return json_response(self, result)

    def master_filter_options(self):
        body = self.parse_body()
        master_id = body.get("master_file_id")
        with connect() as conn:
            master = conn.execute("select * from excel_files where id=?", (master_id,)).fetchone()
        if not master:
            return error_response(self, "請先選用要查詢的總表")
        sheet_name = body.get("master_sheet") or "原物料"
        result = master_filter_options(Path(master["stored_path"]), sheet_name)
        result["file_name"] = master["original_name"]
        result["file_kind"] = master["kind"]
        return json_response(self, result)

    def get_material_candidates(self):
        body = self.parse_body()
        master_id = body.get("master_file_id")
        survey_id = body.get("survey_file_id")
        with connect() as conn:
            master = conn.execute("select * from excel_files where id=?", (master_id,)).fetchone()
            survey = conn.execute("select * from excel_files where id=?", (survey_id,)).fetchone()
        if not master or not survey:
            return error_response(self, "請先上傳總表與調查表")
        master_sheet = body.get("master_sheet") or "原物料"
        supplier_code = clean_text(body.get("supplier_code"))
        category = clean_text(body.get("category") or master_sheet)
        subcategory = clean_text(body.get("subcategory"))
        survey_version = clean_text(body.get("survey_version"))
        material_keyword = clean_text(body.get("material_keyword"))
        _wb, _ws, survey_meta = parse_survey_header(Path(survey["stored_path"]))
        survey_meta.update({"warnings": [], "rows": None})
        search_material_name = material_keyword or survey_meta.get("material_name", "")
        candidates = material_candidates(Path(master["stored_path"]), master_sheet, search_material_name, supplier_code, category)
        return json_response(self, {"survey_meta": survey_meta, "search_material_name": search_material_name, "candidates": candidates})

    def get_material_columns(self):
        body = self.parse_body()
        master_id = body.get("master_file_id")
        with connect() as conn:
            master = conn.execute("select * from excel_files where id=?", (master_id,)).fetchone()
        if not master:
            return error_response(self, "請先選用總表")
        master_sheet = body.get("master_sheet") or "原物料"
        return json_response(self, {"columns": material_columns(Path(master["stored_path"]), master_sheet)})

    def job_results(self, job_id: str, query: dict):
        page = int((query.get("page") or ["1"])[0])
        page_size = int((query.get("page_size") or ["50"])[0])
        diff_type = (query.get("type") or [""])[0]
        offset = (page - 1) * page_size
        where = "where job_id=?"
        params: list = [job_id]
        if not diff_type or diff_type == "all":
            where += " and diff_type in ('added', 'modified')"
        else:
            where += " and diff_type=?"
            params.append(diff_type)
        with connect() as conn:
            total = conn.execute(f"select count(*) c from comparison_results {where}", params).fetchone()["c"]
            rows = conn.execute(
                f"select * from comparison_results {where} order by case diff_type when 'conflict' then 0 when 'modified' then 1 when 'added' then 2 when 'removed' then 3 else 4 end limit ? offset ?",
                [*params, page_size, offset],
            ).fetchall()
        data = []
        for r in rows:
            item = dict(r)
            item["old"] = json.loads(item.pop("old_json"))
            item["new"] = json.loads(item.pop("new_json"))
            item["changes"] = json.loads(item.pop("changes_json"))
            data.append(item)
        return json_response(self, {"total": total, "rows": data, "page": page, "page_size": page_size})

    def confirm(self, job_id: str):
        body = self.parse_body()
        ids = body.get("result_ids") or []
        action = body.get("action") or "accept_new"
        allowed_actions = {"accept_new", "keep_old", "confirm_add", "confirm_remove", "mark_inactive", "ignore"}
        allowed_by_type = {
            "added": {"accept_new", "confirm_add", "ignore"},
            "modified": {"accept_new", "keep_old", "ignore"},
            "removed": {"keep_old", "confirm_remove", "mark_inactive", "ignore"},
            "same": {"ignore"},
        }
        if action not in allowed_actions:
            return error_response(self, "處理方式不支援")
        if not ids:
            return error_response(self, "請至少勾選一筆差異")
        placeholders = ",".join("?" for _ in ids)
        with connect() as conn:
            rows = conn.execute(
                f"select id, diff_type from comparison_results where job_id=? and id in ({placeholders})",
                [job_id, *ids],
            ).fetchall()
            if len(rows) != len(ids):
                return error_response(self, "有勾選資料不屬於目前比對工作，請重新整理後再操作")
            invalid = [row["diff_type"] for row in rows if action not in allowed_by_type.get(row["diff_type"], set())]
            if invalid:
                label_map = {"added": "新增", "modified": "修改", "removed": "刪除", "same": "相同", "conflict": "衝突"}
                invalid_labels = sorted({label_map.get(t, t) for t in invalid})
                return error_response(self, f"「{action}」不能套用在 {', '.join(invalid_labels)} 資料，請改用符合差異類型的確認方式")
            conn.executemany(
                "update comparison_results set action=?, confirmed_at=? where id=? and job_id=?",
                [(action, now(), result_id, job_id) for result_id in ids],
            )
        log("確認", f"{action} {len(ids)} 筆", job_id)
        return json_response(self, {"ok": True, "count": len(ids), "action": action})

    def auto_confirm(self, job_id: str):
        timestamp = now()
        with connect() as conn:
            job = conn.execute("select id from comparison_jobs where id=?", (job_id,)).fetchone()
            if not job:
                return error_response(self, "找不到比對工作", 404)
            added = conn.execute(
                "update comparison_results set action='confirm_add', confirmed_at=? where job_id=? and diff_type='added' and action='pending'",
                (timestamp, job_id),
            ).rowcount
            modified = conn.execute(
                "update comparison_results set action='accept_new', confirmed_at=? where job_id=? and diff_type='modified' and action='pending' and changes_json not like '%同批重複資料%'",
                (timestamp, job_id),
            ).rowcount
        log("一鍵確認", f"新增 {added} 筆、修改 {modified} 筆", job_id)
        return json_response(self, {"ok": True, "added": added, "modified": modified})

    def cancel_job(self, job_id: str):
        with connect() as conn:
            job = conn.execute("select status from comparison_jobs where id=?", (job_id,)).fetchone()
            if not job:
                return error_response(self, "找不到比對工作", 404)
            if job["status"] == "applied":
                return error_response(self, "此比對工作已套用，不能取消。若要重來，請放棄本輪暫存更動版後重新比對。")
            deleted_results = conn.execute("delete from comparison_results where job_id=?", (job_id,)).rowcount
            conn.execute("delete from comparison_jobs where id=?", (job_id,))
        log("取消比對", f"取消比對工作並刪除差異暫存 {deleted_results} 筆", job_id)
        return json_response(self, {"ok": True, "deleted_results": deleted_results})

    def apply(self, job_id: str):
        with connect() as conn:
            job = conn.execute("select status from comparison_jobs where id=?", (job_id,)).fetchone()
            if not job:
                return error_response(self, "找不到比對工作", 404)
            if job["status"] == "file_deleted":
                return error_response(self, "此比對工作引用的檔案已刪除，不能再套用。")
            conflicts = conn.execute("select count(*) c from comparison_results where job_id=? and diff_type='conflict'", (job_id,)).fetchone()["c"]
            pending = conn.execute("select count(*) c from comparison_results where job_id=? and diff_type!='same' and diff_type!='conflict' and action='pending'", (job_id,)).fetchone()["c"]
        if conflicts:
            return error_response(self, "仍有衝突資料，請先處理後再套用")
        if pending:
            return error_response(self, f"仍有 {pending} 筆差異尚未確認，請先確認或暫不處理")
        result = apply_job(job_id)
        with connect() as conn:
            conn.execute("update comparison_jobs set status='applied' where id=?", (job_id,))
        return json_response(self, result)

    def download(self, name: str):
        name = unquote(name)
        path = (OUTPUTS / name).resolve()
        if not str(path).startswith(str(OUTPUTS.resolve())) or not path.exists():
            return error_response(self, "找不到下載檔案", 404)
        self.send_response(200)
        self.send_header("Content-Type", mimetypes.guess_type(path.name)[0] or "application/octet-stream")
        self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(path.name)}")
        self.send_header("Content-Length", str(path.stat().st_size))
        self.end_headers()
        with path.open("rb") as f:
            shutil.copyfileobj(f, self.wfile)

    def download_file(self, file_id: str):
        with connect() as conn:
            row = conn.execute("select * from excel_files where id=?", (file_id,)).fetchone()
        if not row:
            return error_response(self, "找不到原始檔案", 404)
        path = Path(row["stored_path"]).resolve()
        if not path.exists():
            return error_response(self, "原始檔案不存在", 404)
        self.send_response(200)
        self.send_header("Content-Type", mimetypes.guess_type(path.name)[0] or "application/octet-stream")
        self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(row['original_name'])}")
        self.send_header("Content-Length", str(path.stat().st_size))
        self.end_headers()
        with path.open("rb") as f:
            shutil.copyfileobj(f, self.wfile)

    def static(self, path: str):
        if path in {"", "/"}:
            path = "/index.html"
        target = (STATIC / path.lstrip("/")).resolve()
        if not str(target).startswith(str(STATIC.resolve())) or not target.exists():
            return error_response(self, "找不到頁面", 404)
        body = target.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", mimetypes.guess_type(target.name)[0] or "text/plain")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


def main() -> None:
    init_db()
    port = int(os.environ.get("PORT", "8000"))
    server = ThreadingHTTPServer(("127.0.0.1", port), App)
    print(f"Excel 物料版本比對與更新系統已啟動：http://127.0.0.1:{port}", flush=True)
    server.serve_forever()


if __name__ == "__main__":
    main()

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


OUTPUT = Path(r"D:\有害物質\outputs\019fa128-90ab-7740-8505-6b6995079dd2\包材調查表_比對測試用_20260730.xlsx")


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "包材調查表"

    blue = PatternFill("solid", fgColor="9CCBEE")
    teal = PatternFill("solid", fgColor="1F6F8B")
    note_fill = PatternFill("solid", fgColor="F3F8FB")
    thin_black = Side(style="thin", color="000000")
    thin_gray = Side(style="thin", color="B7C3CC")
    header_border = Border(left=thin_black, right=thin_black, top=thin_black, bottom=thin_black)
    data_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    ws.merge_cells("A1:I1")
    ws["A1"] = "包材調查表 - 比對測試用"
    ws["A1"].fill = teal
    ws["A1"].font = Font(name="Microsoft JhengHei", bold=True, color="FFFFFF", size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:I2")
    ws["A2"] = "使用方式：在系統「調查表上傳 > 包材調查表」填入供應商名稱/代號與品項/包材名稱後，選擇本檔加入包材批次。"
    ws.merge_cells("A3:I3")
    ws["A3"] = "建議測試：供應商名稱填「正隆」、品項/包材名稱填「紙箱」可測既有包材修改；也可填新供應商或新品項測右側欄位新增。"
    for row in (2, 3):
        cell = ws.cell(row, 1)
        cell.fill = note_fill
        cell.font = Font(name="Microsoft JhengHei", color="244657")
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    headers = [
        ["No.", "化學物質名稱", "CAS NO.", "等級", "允許濃度", "有 Yes", "含量", "使用用途", "計畫削減"],
        ["", "chemical element", "CAS NO.", "Level", "Threshold Limited(ppm)", "無 No", "content(ppm)", "Purpose", "Hazardous Reduction Plan"],
    ]
    for row_index, row_values in enumerate(headers, 7):
        for col_index, value in enumerate(row_values, 1):
            cell = ws.cell(row_index, col_index, value)
            cell.fill = blue
            cell.font = Font(name="Microsoft JhengHei", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = header_border

    rows = [
        [1, "Cadmium ( Cd ) / Cadmium Compounds(鎘及化合物)", "7440-43-9", 1, "ND", "無 No", "", "", ""],
        [2, "Lead ( Pb ) / Lead Compounds(鉛及化合物)", "7439-92-1、附表8.1.1", 1, "0.005", "有 Yes", "紙材: 0.005", "Raw material residue", ""],
        [3, "Mercury ( Hg ) / Mercury Compounds(汞及化合物)", "7439-97-6", 1, "ND", "無 No", "", "", ""],
        [4, "Hexavalent-Chromium(Cr6+)Compounds(六價鉻化合物)", "18540-29-9", 1, "ND", "無 No", "", "", ""],
        [5, "PBBs (Polybrominated biphenyls)聚溴聯苯", "059536-65-1", 1, "ND", "無 No", "", "", ""],
        [6, "Polybrominated diphenyl ethers(PBDEs)聚溴二苯醚", "1163-19-5", 1, "ND", "無 No", "", "", ""],
        [7, "氯(Cl) Chlorine", "22537-15-1", 2, "97", "有 Yes", "面紙: 97", "Paper additive", ""],
        [8, "溴(Br) Bromine", "10097-32-2", 2, "ND", "無 No", "", "", ""],
        [9, "Polyvinyl chloride(PVC)and PVC blends聚氯乙烯以及聚氯乙烯混合物", "9002-86-2", 2, "ND", "無 No", "", "", ""],
        [10, "Perfluoroocatane sulfonate(PFOS)全氟辛烷磺酸", "1763-23-1", 2, "ND", "無 No", "", "", ""],
        [11, "Formaldehyde 甲醛", "50-00-0", 2, "<75", "有 Yes", "膠水殘留: 12", "Adhesive residue", ""],
        [12, "測試新增物質 Test Added Substance", "99999-99-9", 2, "0.010", "有 Yes", "0.010", "新增測試用", ""],
    ]
    for row_index, row_values in enumerate(rows, 9):
        for col_index, value in enumerate(row_values, 1):
            cell = ws.cell(row_index, col_index, value)
            cell.font = Font(name="Microsoft JhengHei")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = data_border

    widths = [7, 48, 26, 10, 20, 14, 22, 22, 28]
    for col_index, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col_index)].width = width
    for row_index in range(1, 21):
        ws.row_dimensions[row_index].height = 24
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[7].height = 32
    ws.row_dimensions[8].height = 32
    ws.freeze_panes = "A9"

    wb.save(OUTPUT)
    wb.close()
    print(OUTPUT)


if __name__ == "__main__":
    main()

from __future__ import annotations

import json
import mimetypes
import os
import shutil
import sqlite3
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path


BASE = os.environ.get("MATERIAL_SYSTEM_BASE", "http://127.0.0.1:8000")
ROOT = Path(r"D:\有害物質\material_system")
DATA = ROOT / "data"
DB = DATA / "material_versions.sqlite3"
MASTER_ID = ""
MASTER_PATH = ROOT.parent / "QP-QA-43-08原物料&包材物質清單2026.xlsx"
RAW_PATH = DATA / "test_tmp" / "raw_test.xlsx"
PKG_PATH = DATA / "test_tmp" / "pkg_test.xlsx"
PKG_NAME_MASTER_PATH = DATA / "test_tmp" / "pkg_name_master.xlsx"
BACKUP_DB = DATA / "test_tmp" / "material_versions.before_e2e.sqlite3"
BASELINE_UPLOADS: set[Path] = set()
BASELINE_OUTPUTS: set[Path] = set()


def create_test_files() -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    RAW_PATH.parent.mkdir(parents=True, exist_ok=True)
    blue = PatternFill("solid", fgColor="9CCBEE")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def build(path: Path, rows: list[list]) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Survey"
        headers = ["No.", "Chemical substance", "CAS NO.", "Level", "Threshold Limited(ppm)", "Yes No", "content(ppm)", "Purpose", "Hazardous Reduction Plan"]
        for row_index in [7, 8]:
            for col, value in enumerate(headers, 1):
                cell = ws.cell(row_index, col, value)
                cell.fill = blue
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
        for row_index, row in enumerate(rows, 9):
            for col, value in enumerate(row, 1):
                cell = ws.cell(row_index, col, value)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col, width in enumerate([8, 42, 26, 10, 22, 14, 20, 18, 26], 1):
            ws.column_dimensions[chr(64 + col)].width = width
        wb.save(path)
        wb.close()

    build(RAW_PATH, [
        [1, "Lead and its compounds(Pb)", "7439-92-1, Appendix 8.1.1", 1, "ND", "No", "", "", ""],
        [2, "CODT", "68783-78-8, 107-64-2", 2, "ND", "Yes", "Toluene: 120", "Solvent", ""],
    ])
    build(PKG_PATH, [
        [1, "Cadmium Compounds", "7440-43-9", 1, "0.005", "Yes", "", "", ""],
        [2, "Lead Compounds", "7439-92-1", 1, "0.005", "Yes", "", "", ""],
        [3, "Mercury Compounds", "7439-97-6", 1, "ND", "No", "", "", ""],
    ])

    wb = Workbook()
    ws = wb.active
    ws.title = "包材"
    ws.cell(3, 4, "供應商")
    ws.cell(3, 5, "包材供應商A")
    headers = ["Item", "化學物質\nChemical substances Group", "允許濃度(ppm)\nAcceptable concentration", "等級\nLevel", "測試膜"]
    for col, value in enumerate(headers, 1):
        cell = ws.cell(4, col, value)
        cell.fill = blue
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row_index, row in enumerate([
        [1, "Cadmium Compounds", "7440-43-9", 1, "No contain"],
        [2, "Lead Compounds", "7439-92-1", 1, "No contain"],
        [3, "Mercury Compounds", "7439-97-6", 1, "No contain"],
    ], 5):
        for col, value in enumerate(row, 1):
            cell = ws.cell(row_index, col, value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(PKG_NAME_MASTER_PATH)
    wb.close()


def request(method: str, path: str, body: dict | None = None, timeout: int = 120) -> dict:
    data = None
    headers = {}
    if body is not None:
        data = json.dumps(body, ensure_ascii=False).encode("utf-8")
        headers["Content-Type"] = "application/json; charset=utf-8"
    req = urllib.request.Request(BASE + path, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as res:
            return json.loads(res.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        raise RuntimeError(f"{method} {path} failed: {exc.read().decode('utf-8', errors='replace')}") from exc


def download(path: str) -> int:
    with urllib.request.urlopen(BASE + path, timeout=120) as res:
        return len(res.read())


def upload(path: Path, kind: str = "survey") -> dict:
    boundary = "----codex-e2e-boundary"
    mime = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    payload = bytearray()
    payload.extend(f"--{boundary}\r\n".encode())
    payload.extend(b'Content-Disposition: form-data; name="kind"\r\n\r\n')
    payload.extend(kind.encode())
    payload.extend(b"\r\n")
    payload.extend(f"--{boundary}\r\n".encode())
    payload.extend(f'Content-Disposition: form-data; name="file"; filename="{path.name}"\r\n'.encode("utf-8"))
    payload.extend(f"Content-Type: {mime}\r\n\r\n".encode())
    payload.extend(path.read_bytes())
    payload.extend(b"\r\n")
    payload.extend(f"--{boundary}--\r\n".encode())
    req = urllib.request.Request(
        BASE + "/api/upload",
        data=bytes(payload),
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=120) as res:
            return json.loads(res.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        raise RuntimeError(f"upload failed: {exc.read().decode('utf-8', errors='replace')}") from exc


def ensure_master() -> dict:
    global MASTER_ID
    dashboard = request("GET", "/api/dashboard")
    latest = (dashboard.get("version_state") or {}).get("latest_master")
    if latest and latest.get("id"):
        MASTER_ID = latest["id"]
        return dashboard
    if not MASTER_PATH.exists():
        raise RuntimeError(f"missing master workbook: {MASTER_PATH}")
    uploaded = upload(MASTER_PATH, "master")
    MASTER_ID = uploaded["id"]
    request("POST", f"/api/files/{MASTER_ID}/set-latest")
    return request("GET", "/api/dashboard")


def check(name: str, ok: bool, detail: str = "") -> str:
    if not ok:
        raise AssertionError(f"{name} failed {detail}")
    return f"{name} OK {detail}".strip()


def snapshot_runtime_files(folder: Path) -> set[Path]:
    if not folder.exists():
        return set()
    return {path.resolve() for path in folder.iterdir() if path.is_file()}


def backup_state() -> None:
    global BASELINE_UPLOADS, BASELINE_OUTPUTS
    RAW_PATH.parent.mkdir(parents=True, exist_ok=True)
    if DB.exists():
        shutil.copy2(DB, BACKUP_DB)
    BASELINE_UPLOADS = snapshot_runtime_files(DATA / "uploads")
    BASELINE_OUTPUTS = snapshot_runtime_files(DATA / "outputs")


def cleanup() -> None:
    if BACKUP_DB.exists():
        shutil.copy2(BACKUP_DB, DB)
    for folder, baseline in [(DATA / "uploads", BASELINE_UPLOADS), (DATA / "outputs", BASELINE_OUTPUTS)]:
        if folder.exists():
            for path in folder.iterdir():
                if path.is_file() and path.resolve() not in baseline:
                    path.unlink()
    for path in [RAW_PATH, PKG_PATH, PKG_NAME_MASTER_PATH, BACKUP_DB]:
        if path.exists():
            path.unlink()


def main() -> None:
    checks: list[str] = []
    try:
        backup_state()
        create_test_files()
        dashboard = request("GET", "/api/dashboard")
        checks.append(check("dashboard", dashboard.get("app_version") == "v1.0.27", dashboard.get("app_version", "")))
        dashboard = ensure_master()

        options = request("POST", "/api/master-filter-options", {"master_file_id": MASTER_ID, "master_sheet": "原物料"})
        checks.append(check("master-filter-options", len(options.get("materials", [])) > 0, f"materials={len(options.get('materials', []))}"))

        search = request("POST", "/api/master-search", {"master_file_id": MASTER_ID, "master_sheet": "包材", "substance": "Lead", "include_empty": True})
        checks.append(check("master-search", len(search.get("rows", [])) > 0, f"rows={len(search.get('rows', []))}"))

        raw_upload = upload(RAW_PATH)
        checks.append(check("raw-upload", bool(raw_upload.get("id")), raw_upload.get("name", "")))
        pkg_upload = upload(PKG_PATH)
        checks.append(check("package-upload", bool(pkg_upload.get("id")), pkg_upload.get("name", "")))
        pkg_name_master_upload = upload(PKG_NAME_MASTER_PATH, "master")
        checks.append(check("package-name-master-upload", bool(pkg_name_master_upload.get("id")), pkg_name_master_upload.get("name", "")))

        pkg_name_compare = request(
            "POST",
            "/api/compare",
            {
                "master_file_id": pkg_name_master_upload["id"],
                "master_sheet": "包材",
                "survey_items": [
                    {
                        "survey_file_id": pkg_upload["id"],
                        "type": "包材",
                        "supplier_code": "包材供應商A",
                        "material_name": "測試膜",
                        "subcategory": "測試",
                        "original_name": pkg_upload["name"],
                    }
                ],
                "survey_version": "E2E-PKG-NAME",
            },
        )
        checks.append(check("package-name-compare", pkg_name_compare["summary"]["old_count"] == 3 and pkg_name_compare["summary"]["new_count"] == 3, json.dumps(pkg_name_compare["summary"], ensure_ascii=False)))
        pkg_name_job = pkg_name_compare["job_id"]
        request("POST", f"/api/jobs/{pkg_name_job}/auto-confirm")
        pkg_name_apply = request("POST", f"/api/jobs/{pkg_name_job}/apply")
        checks.append(check("package-name-apply", pkg_name_apply.get("changed", {}).get("columns_added", 0) == 0 and pkg_name_apply.get("changed", {}).get("content_updated", 0) >= 1, json.dumps(pkg_name_apply.get("changed", {}), ensure_ascii=False)))

        cancel_compare = request(
            "POST",
            "/api/compare",
            {
                "master_file_id": MASTER_ID,
                "master_sheet": "原物料",
                "survey_items": [
                    {
                        "survey_file_id": raw_upload["id"],
                        "type": "原物料",
                        "supplier_code": "RAWTEST001",
                        "material_name": "RAW-TEST",
                        "subcategory": "測試",
                        "original_name": raw_upload["name"],
                    }
                ],
                "survey_version": "E2E-CANCEL",
            },
        )
        cancel_job = cancel_compare["job_id"]
        cancel = request("POST", f"/api/jobs/{cancel_job}/cancel")
        checks.append(check("cancel-compare", cancel.get("ok") is True and cancel.get("deleted_results", 0) >= 1, json.dumps(cancel, ensure_ascii=False)))

        raw_compare = request(
            "POST",
            "/api/compare",
            {
                "master_file_id": MASTER_ID,
                "master_sheet": "原物料",
                "survey_items": [
                    {
                        "survey_file_id": raw_upload["id"],
                        "type": "原物料",
                        "supplier_code": "RAWTEST001",
                        "material_name": "RAW-TEST",
                        "subcategory": "測試",
                        "original_name": raw_upload["name"],
                    }
                ],
                "survey_version": "E2E-RAW",
            },
        )
        checks.append(check("raw-compare", raw_compare["summary"]["new_count"] == 2, json.dumps(raw_compare["summary"], ensure_ascii=False)))
        raw_job = raw_compare["job_id"]
        raw_results = request("GET", f"/api/jobs/{raw_job}/results?type=all&page=1&page_size=50")
        checks.append(check("raw-results", raw_results.get("total", 0) >= 1, f"total={raw_results.get('total', 0)}"))
        raw_auto = request("POST", f"/api/jobs/{raw_job}/auto-confirm")
        checks.append(check("raw-auto-confirm", raw_auto.get("added", 0) + raw_auto.get("modified", 0) >= 1, json.dumps(raw_auto, ensure_ascii=False)))
        raw_apply = request("POST", f"/api/jobs/{raw_job}/apply")
        checks.append(check("raw-apply", bool(raw_apply.get("updated_file_id")), json.dumps(raw_apply.get("changed", {}), ensure_ascii=False)))
        checks.append(check("download-updated", download("/download/" + urllib.parse.quote(raw_apply["updated_file"])) > 1000))

        pkg_compare = request(
            "POST",
            "/api/compare",
            {
                "master_file_id": raw_apply["updated_file_id"],
                "master_sheet": "包材",
                "survey_items": [
                    {
                        "survey_file_id": pkg_upload["id"],
                        "type": "包材",
                        "supplier_code": "PKGTEST001",
                        "material_name": "PKG-TEST",
                        "subcategory": "測試",
                        "original_name": pkg_upload["name"],
                    }
                ],
                "survey_version": "E2E-PKG",
            },
        )
        checks.append(check("package-compare", pkg_compare["summary"]["old_count"] >= 70 and pkg_compare["summary"]["new_count"] == 3, json.dumps(pkg_compare["summary"], ensure_ascii=False)))
        pkg_job = pkg_compare["job_id"]
        pkg_auto = request("POST", f"/api/jobs/{pkg_job}/auto-confirm")
        checks.append(check("package-auto-confirm", pkg_auto.get("added", 0) + pkg_auto.get("modified", 0) >= 1, json.dumps(pkg_auto, ensure_ascii=False)))
        pkg_apply = request("POST", f"/api/jobs/{pkg_job}/apply")
        checks.append(check("package-apply", bool(pkg_apply.get("updated_file_id")) and pkg_apply.get("changed", {}).get("columns_added", 0) >= 1, json.dumps(pkg_apply.get("changed", {}), ensure_ascii=False)))

        finalize = request("POST", "/api/finalize-working-version")
        checks.append(check("finalize", finalize.get("ok") is True, finalize.get("message", "")))
        deleted = request("DELETE", f"/api/files/{raw_upload['id']}")
        checks.append(check("delete-survey", deleted.get("ok") is True, deleted.get("deleted_id", "")))
        checks.append(check("download-master", download(f"/download-file/{MASTER_ID}") > 1000))
    finally:
        cleanup()

    for item in checks:
        print(item)


if __name__ == "__main__":
    main()
